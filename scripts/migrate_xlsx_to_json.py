"""One-time migration: South_Wales_Walks_Database.xlsx -> walks/*.json.

Run this once to seed the walks/ directory from the existing spreadsheet.
After migration the JSON files become the source of truth; build_walks.py
reads them back to regenerate the xlsx.

Usage:
    python scripts/migrate_xlsx_to_json.py            # write walks/*.json
    python scripts/migrate_xlsx_to_json.py --dry-run  # print sample, no writes

Idempotent: re-running overwrites files with the same content.
"""
from __future__ import annotations
import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

HERE  = Path(__file__).resolve().parent.parent
XLSX  = HERE / "South_Wales_Walks_Database.xlsx"
WALKS = HERE / "walks"

# xlsx header -> JSON field. Anything not in this map is dropped (e.g. "Map Link"
# is derived from postcode at build time, "Distance (km)" is derived from mi).
FIELD_MAP = {
    "ID":                         "id",
    "Walk Name":                  "name",
    "Region":                     "region",
    "Sub-area":                   "sub_area",
    "Nearest Town":               "nearest_town",
    "Distance (mi)":              "distance_mi",
    "Elevation Gain (m)":         "elevation_gain_m",
    "Est. Time (hrs)":            "est_time_hrs",
    "Difficulty":                 "difficulty",
    "Route Type":                 "route_type",
    "Terrain":                    "terrain",
    "Dogs Allowed":               "dogs_allowed",
    "Dog Lead Policy":            "dog_lead_policy",
    "Pushchair Friendly":         "pushchair_friendly",
    "Waymarked":                  "waymarked",
    "Best Season":                "best_season",
    "Key Features / Highlights":  "highlights",
    "Points of Interest":         "points_of_interest",
    "Viewpoints & Beauty Spots":  "viewpoints",
    "Water Features":             "water_features",
    "Picnic Spots":               "picnic_spots",
    "Parking & Start":            "parking_start",
    "Food & Drink Nearby":        "food_drink_nearby",
    "Toilets":                    "toilets",
    "Public Transport":           "public_transport",
    "Hazards / Notes":            "hazards_notes",
    "Start Postcode":             "start_postcode",
    "Drive from Monmouth (mins)": "drive_from_monmouth_mins",
}

# JSON field order — mirrors the schema and gives readable diffs.
KEY_ORDER = [
    "id", "slug", "name",
    "region", "sub_area", "nearest_town",
    "distance_mi", "elevation_gain_m", "est_time_hrs",
    "difficulty", "route_type", "terrain",
    "dogs_allowed", "dog_lead_policy", "pushchair_friendly", "waymarked",
    "best_season",
    "highlights", "points_of_interest", "viewpoints",
    "water_features", "picnic_spots",
    "parking_start", "food_drink_nearby", "toilets",
    "public_transport", "hazards_notes",
    "start_postcode", "drive_from_monmouth_mins",
]

INT_FIELDS   = {"id", "elevation_gain_m", "drive_from_monmouth_mins"}
FLOAT_FIELDS = {"distance_mi", "est_time_hrs"}


def slugify(name: str) -> str:
    """URL-safe kebab-case slug. Mirrors the pattern in schema."""
    s = name.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def read_walks_from_xlsx() -> list[dict]:
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Walks"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        raw = {headers[c]: ws.cell(r, c + 1).value for c in range(len(headers))}
        rec: dict = {}
        for src, dst in FIELD_MAP.items():
            v = raw.get(src)
            if v is None or (isinstance(v, str) and not v.strip()):
                rec[dst] = None
                continue
            if dst in INT_FIELDS and isinstance(v, (int, float)):
                rec[dst] = int(v)
            elif dst in FLOAT_FIELDS and isinstance(v, (int, float)):
                rec[dst] = round(float(v), 2)
            elif isinstance(v, str):
                rec[dst] = v.strip()
            else:
                rec[dst] = v
        if not rec.get("id") or not rec.get("name"):
            continue
        # Uppercase postcodes, drop internal double-spaces
        if rec.get("start_postcode"):
            rec["start_postcode"] = re.sub(r"\s+", " ", rec["start_postcode"]).upper()
        rec["slug"] = slugify(rec["name"])
        rows.append(rec)
    return rows


def ordered(rec: dict) -> dict:
    """Return a dict in KEY_ORDER so JSON files diff cleanly."""
    return {k: rec.get(k) for k in KEY_ORDER if k in rec}


def filename_for(rec: dict) -> str:
    return f"{rec['id']:03d}-{rec['slug']}.json"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not XLSX.exists():
        print(f"ERROR: xlsx not found at {XLSX}", file=sys.stderr)
        sys.exit(1)

    walks = read_walks_from_xlsx()
    print(f"Read {len(walks)} walks from {XLSX.name}")

    if args.dry_run:
        sample = walks[0]
        print("\n--- sample JSON (walk 1) ---")
        print(json.dumps(ordered(sample), indent=2, ensure_ascii=False))
        print(f"\n--- would write {len(walks)} files to {WALKS}/ ---")
        return

    WALKS.mkdir(exist_ok=True)
    # Remove stale JSONs that no longer match any walk
    expected = {filename_for(w) for w in walks}
    for existing in WALKS.glob("*.json"):
        if existing.name not in expected:
            print(f"  - removing stale: {existing.name}")
            existing.unlink()

    written = 0
    for w in walks:
        path = WALKS / filename_for(w)
        payload = json.dumps(ordered(w), indent=2, ensure_ascii=False) + "\n"
        if path.exists() and path.read_text(encoding="utf-8") == payload:
            continue
        path.write_text(payload, encoding="utf-8")
        written += 1
    print(f"Wrote {written} / {len(walks)} files to {WALKS}/")


if __name__ == "__main__":
    main()
