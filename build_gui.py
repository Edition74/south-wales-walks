"""Build a single-file HTML GUI for the South Wales Walks Database.

Reads the existing xlsx, auto-tags each walk with points-of-interest categories
derived from its features/POI/water/terrain text, and emits a responsive,
mobile-first web app in an editorial/outdoor-tourism visual style:

  - Moody forest green + cream + bracken orange palette
  - Fraunces display serif + Inter body sans
  - Hero photo, stats strip, featured walks, region tiles, finder
  - Per-region accent colour + Unsplash CDN imagery (with gradient fallbacks)
"""
import datetime
import html
import json
import os
import re
from pathlib import Path
from openpyxl import load_workbook

HERE = Path(__file__).parent
XLSX = HERE / "South_Wales_Walks_Database.xlsx"
OUT  = HERE / "index.html"
PHOTOS_CACHE = HERE / "photos_cache.json"

# Ratings backend — injected into the client bundle. Safe to be public: the
# anon key is designed for browsers (row-level security is what keeps the
# data safe). Set as GitHub Actions secrets for the production build.
SUPABASE_URL      = os.environ.get("SUPABASE_URL",      "")
SUPABASE_ANON_KEY = os.environ.get("SUPABASE_ANON_KEY", "")

# Optional: real Geograph photo cache keyed by walk ID. Populated by fetch_photos.py.
# When present and a walk has photos, they replace the generic Unsplash gallery.
try:
    _photo_cache = json.loads(PHOTOS_CACHE.read_text(encoding="utf-8")).get("walks", {}) \
        if PHOTOS_CACHE.exists() else {}
except Exception:
    _photo_cache = {}
print(f"Loaded {sum(1 for v in _photo_cache.values() if v.get('photos'))} walks with Geograph photos")

wb = load_workbook(XLSX, data_only=True)
ws = wb["Walks"]

# Column index map (1-based)
COLS = {}
for c in range(1, ws.max_column + 1):
    COLS[ws.cell(1, c).value] = c

# ---------------------------------------------------------------------------
# Tagging: derive POI tags from keywords in free-text fields.
# ---------------------------------------------------------------------------
TAG_RULES = [
    ("Castle",               [r"\bcastle\b", r"\bfort\b", r"hillfort", r"ramparts"]),
    ("Waterfall",            [r"waterfall", r"\bfall(s)?\b", r"cascade", r"sgwd"]),
    ("Beach / Coastal",      [r"\bbeach\b", r"\bbay\b", r"\bcove\b", r"coast path",
                              r"cliff", r"\bshore", r"headland", r"lighthouse"]),
    ("Lake / Tarn / Pond",   [r"\blake\b", r"\bpond\b", r"\btarn\b", r"reservoir",
                              r"\bllyn\b", r"lily pond", r"lily ponds"]),
    ("Mountain / Summit",    [r"summit", r"peak", r"\bridge\b", r"\bfan\b",
                              r"mountain", r"mynydd", r"\btrig\b"]),
    ("Prehistoric / Roman",  [r"iron age", r"bronze age", r"neolithic", r"roman",
                              r"hillfort", r"burial", r"\bcairn", r"stone circle",
                              r"standing stone", r"quoit", r"dolmen"]),
    ("Abbey / Church",       [r"\babbey\b", r"priory", r"chapel", r"\bchurch\b",
                              r"monaster", r"temple", r"hermit"]),
    ("Woodland / Forest",    [r"woodland", r"\bforest\b", r"\bwood(s)?\b", r"coppice",
                              r"ancient oak", r"bluebell"]),
    ("Wildlife / Nature Reserve",
                             [r"\brspb\b", r"\bnnr\b", r"\bsssi\b", r"nature reserve",
                              r"wetland", r"salt ?marsh", r"bird", r"otter",
                              r"peregrine", r"seal"]),
    ("Industrial Heritage",  [r"tramroad", r"tramway", r"\bcanal\b", r"\bmine\b",
                              r"\bcollier", r"ironworks", r"quarr", r"\bwharf\b",
                              r"industrial", r"pit\b", r"railway", r"viaduct",
                              r"tidal mill", r"big pit", r"locks"]),
    ("River / Gorge",        [r"\briver\b", r"riverside", r"\bgorge\b", r"\bafon\b",
                              r"\busk\b", r"\bwye\b", r"\bmonnow\b", r"\btrothy\b"]),
    ("Waymarked Long-Distance Path",
                             [r"offa'?s dyke", r"wales coast path", r"wye valley walk",
                              r"taff trail", r"three castles walk", r"ncn"]),
    ("Pub on Route",         []),
    ("Family Friendly",      []),
    ("Dog Friendly Pub",     []),
]

def tag_walk(row):
    blob = " ".join(str(row.get(k, "")).lower()
                    for k in ("Walk Name", "Sub-area", "Terrain",
                              "Key Features / Highlights",
                              "Points of Interest", "Viewpoints & Beauty Spots",
                              "Water Features", "Hazards / Notes", "Waymarked"))
    tags = []
    for name, patterns in TAG_RULES:
        if not patterns:
            continue
        for p in patterns:
            if re.search(p, blob):
                tags.append(name)
                break
    food = str(row.get("Food & Drink Nearby", "")).lower()
    if any(w in food for w in (" inn", " pub", " arms", " head", "tavern",
                               "hotel", "tea room", "tearoom", "cafe", "café")):
        tags.append("Pub / Cafe on Route")
    if "dog friendly" in food:
        tags.append("Dog-Friendly Pub")
    diff = str(row.get("Difficulty", "")).lower()
    pushchair = str(row.get("Pushchair Friendly", "")).lower()
    if diff.startswith("easy") and pushchair in ("yes", "partial"):
        tags.append("Family Friendly")
    return sorted(set(tags))


# ---------------------------------------------------------------------------
# Build walk records
# ---------------------------------------------------------------------------
walks = []
for r in range(2, ws.max_row + 1):
    row = {h: ws.cell(r, c).value for h, c in COLS.items()}
    row["tags"] = tag_walk(row)
    miles = row.get("Distance (mi)") or 0
    row["Distance (km)"] = round(float(miles) * 1.609344, 1) if miles else 0
    row["Drive from Monmouth (mins)"] = row.get("Drive from Monmouth (mins)") or 999
    walks.append(row)

def short(row):
    return {
        "id": row["ID"],
        "name": row["Walk Name"],
        "region": row["Region"],
        "sub": row["Sub-area"],
        "town": row["Nearest Town"],
        "miles": row["Distance (mi)"],
        "km": row["Distance (km)"],
        "elev": row["Elevation Gain (m)"],
        "time": row["Est. Time (hrs)"],
        "difficulty": row["Difficulty"],
        "route": row["Route Type"],
        "terrain": row["Terrain"],
        "dogs": row["Dogs Allowed"],
        "leash": row["Dog Lead Policy"],
        "pushchair": row["Pushchair Friendly"],
        "waymarked": row["Waymarked"],
        "season": row["Best Season"],
        "features": row["Key Features / Highlights"],
        "poi": row["Points of Interest"],
        "views": row["Viewpoints & Beauty Spots"],
        "water": row["Water Features"],
        "picnic": row["Picnic Spots"],
        "parking": row["Parking & Start"],
        "food": row["Food & Drink Nearby"],
        "toilets": row["Toilets"],
        "transport": row["Public Transport"],
        "notes": row["Hazards / Notes"],
        "postcode": row["Start Postcode"],
        "drive": row["Drive from Monmouth (mins)"],
        "tags": row["tags"],
        "images": [],  # filled in below once REGION_META/PHOTO_BANK are defined
        "photo_credits": [],  # filled in below if Geograph cache has entries
    }

data = [short(w) for w in walks]
print(f"Prepared {len(data)} walks")

# ---------------------------------------------------------------------------
# Region metadata: short name, tagline, accent colour, Unsplash image.
# Images are referenced by their stable Unsplash photo-id path. If a photo
# ever 404s, the CSS gradient fallback underneath still looks good.
# ---------------------------------------------------------------------------
UNSPLASH = "https://images.unsplash.com"
def u(photo_id, w=1200, q=70):
    return f"{UNSPLASH}/{photo_id}?auto=format&fit=crop&w={w}&q={q}"

REGION_META = {
    "Brecon Beacons / Bannau Brycheiniog": {
        "short": "Brecon Beacons",
        "tagline": "Twin summits, glacial tarns and waterfall country.",
        "image": u("photo-1464822759023-fed622ff2c3b"),
        "accent": "#3d5a3b",
    },
    "Gower & Swansea Bay": {
        "short": "Gower & Swansea Bay",
        "tagline": "Britain's first AONB — limestone cliffs, tidal bays, Worm's Head.",
        "image": u("photo-1507525428034-b723cf961d3e"),
        "accent": "#2c5470",
    },
    "Pembrokeshire (South)": {
        "short": "South Pembrokeshire",
        "tagline": "Coast path icons: lily ponds, sea stacks and the Green Bridge.",
        "image": u("photo-1476514525535-07fb3b4ae5f1"),
        "accent": "#1f6d6a",
    },
    "Valleys & Vale of Glamorgan": {
        "short": "Valleys & Vale",
        "tagline": "Industrial heritage, heritage coast and green lungs above Cardiff.",
        "image": u("photo-1500530855697-b586d89ba3ee"),
        "accent": "#5a4a2a",
    },
    "Wye Valley & Monmouthshire": {
        "short": "Wye Valley & Monmouthshire",
        "tagline": "Hanging woods, river meanders and Norman castles on every other hill.",
        "image": u("photo-1441974231531-c6227db76b6e"),
        # Wood-brown to differentiate from Brecon Beacons' moss green and
        # echo the hanging-woods/sandstone character of the lower Wye.
        "accent": "#6b4a2a",
    },
    "Mid Wales (Powys & Ceredigion)": {
        "short": "Mid Wales",
        "tagline": "Elan Valley dams, the Cambrian wilderness and Cardigan Bay cliffs.",
        "image": u("photo-1469474968028-56623f02e42e"),
        # Olive moss, to slot between Brecon Beacons and the warmer south
        # accents instead of the previous off-theme cool navy.
        "accent": "#5d6b3a",
    },
    "Carmarthenshire & West Wales": {
        "short": "Carmarthenshire",
        "tagline": "Tywi valley meadows, Dinefwr deer and Cenarth's salmon leap.",
        "image": u("photo-1449034446853-66c86144b0ad"),
        "accent": "#6b5a26",
    },
    "English Borders (Forest of Dean & Herefordshire)": {
        "short": "Borders & Forest of Dean",
        "tagline": "Scowles, sculptures and the Cat's Back above the Olchon valley.",
        "image": u("photo-1448375240586-882707db888b"),
        "accent": "#3f3a25",
    },
}

# ---------------------------------------------------------------------------
# Interactive Wales map: hand-drawn region silhouettes. The goal is instant
# regional identity ("that's the Gower peninsula") rather than cartographic
# precision. Each region is a clickable zone that toggles the matching
# filter chip below. North Wales is rendered greyed-out as "coming soon",
# and the English Borders zone extends east of the main Wales outline with
# a dashed border to signal that it's over-the-line.
# ---------------------------------------------------------------------------
WALES_MAP_VIEWBOX = "0 0 820 780"

# Real Wales boundary, derived from ONS Census 2021 country boundary
# (wales_ctry_2022.kml). Simplified via Ramer–Douglas–Peucker (epsilon=1.2)
# and projected to the 820x780 viewBox using equirectangular projection
# (lon -5.45..-2.55, lat 51.30..53.48). See parse_wales_kml.py.
WALES_MAINLAND_PATH = (
    "M 213.9,647.4 L 208.9,654.2 L 196.6,653.7 L 195.1,655.4 L 196.3,656.9 L 191.2,657.4 L 189.2,660.2 L 175.1,655.9 L 167.1,656.3 L 166.2,659.1 L 156.7,662.5 L 154.7,666.5 L 156.6,669.1 L 151.7,668.9 L 148.1,674.1 L 141.8,673.7 L 130.8,668.3 L 125.2,669.7 L 110.4,665.3 L 109.7,662.2 L 112.2,659.8 L 110.3,659.3 L 112.2,658.4 L 107.3,650.0 L 98.2,648.6 L 95.2,646.3 L 92.2,647.2 L 93.9,645.9 L 91.6,643.0 L 95.1,642.3 L 94.9,639.7 L 96.3,639.6 L 104.6,640.5 L 106.5,643.6 L 110.6,644.2 L 112.7,639.1 L 119.4,638.6 L 132.6,640.6 L 131.1,642.6 L 134.2,643.7 L 133.5,638.8 L 137.9,636.4 L 146.4,634.4 L 150.3,636.7 L 153.8,634.0 L 158.0,634.7 L 159.3,630.9 L 163.4,630.6 L 159.8,630.1 L 160.3,624.7 L 156.1,622.7 L 159.5,611.5 L 157.0,612.9 L 151.9,611.3 L 156.9,615.0 L 154.8,622.0 L 159.2,626.0 L 156.8,633.3 L 153.6,632.9 L 149.9,635.1 L 145.1,633.9 L 131.9,637.2 L 117.5,633.0 L 110.0,634.0 L 110.2,636.7 L 112.0,636.4 L 109.2,637.1 L 109.4,634.3 L 102.8,634.6 L 99.6,629.7 L 94.5,633.2 L 91.1,631.1 L 89.2,632.4 L 81.5,631.5 L 79.9,634.3 L 85.0,636.0 L 82.9,636.4 L 82.9,641.2 L 79.5,641.4 L 79.5,643.5 L 77.4,643.9 L 73.8,640.3 L 73.7,633.8 L 70.0,633.2 L 69.8,630.9 L 66.8,629.1 L 60.8,630.6 L 62.4,628.7 L 55.6,625.0 L 55.3,623.2 L 67.3,624.6 L 69.3,622.7 L 69.7,617.0 L 74.4,617.3 L 81.9,610.8 L 94.1,612.1 L 97.5,604.5 L 94.6,588.5 L 87.7,578.3 L 75.8,579.3 L 75.7,577.0 L 67.3,577.2 L 65.7,575.1 L 61.9,574.9 L 57.6,575.1 L 56.1,577.1 L 55.1,575.5 L 47.4,576.0 L 47.9,577.6 L 43.7,580.3 L 43.1,577.8 L 37.8,579.8 L 36.7,578.0 L 40.3,574.3 L 37.8,570.8 L 41.1,569.8 L 42.9,566.8 L 41.2,564.3 L 38.3,564.5 L 45.5,560.2 L 54.0,560.0 L 60.4,554.8 L 67.8,553.5 L 67.0,551.5 L 70.5,550.6 L 69.5,549.1 L 71.7,547.2 L 82.7,548.1 L 86.8,545.3 L 85.2,543.5 L 87.8,542.1 L 92.8,543.6 L 96.8,543.2 L 98.9,540.1 L 102.8,540.8 L 103.9,539.0 L 102.4,536.7 L 104.2,534.6 L 101.4,534.2 L 102.2,533.0 L 100.2,531.8 L 106.6,527.9 L 101.9,524.1 L 106.9,520.1 L 106.1,518.5 L 120.6,520.5 L 121.4,522.2 L 129.6,520.2 L 130.9,523.7 L 135.8,525.0 L 132.2,524.3 L 129.4,527.8 L 132.7,527.7 L 131.1,529.3 L 135.0,530.7 L 135.6,528.9 L 149.7,525.8 L 152.4,522.0 L 150.2,520.2 L 152.5,517.1 L 157.2,518.3 L 157.2,521.0 L 159.6,522.1 L 169.6,521.6 L 172.7,518.1 L 170.3,515.3 L 173.0,511.6 L 186.1,508.5 L 193.1,501.8 L 195.3,501.7 L 197.2,496.3 L 200.1,494.3 L 198.1,492.2 L 203.2,487.3 L 209.2,490.8 L 214.4,489.3 L 214.9,491.1 L 215.9,483.0 L 227.4,481.7 L 228.2,480.1 L 245.2,480.2 L 249.6,477.8 L 255.4,481.1 L 263.1,480.9 L 276.3,472.0 L 278.2,467.3 L 278.5,469.0 L 281.4,469.5 L 288.7,465.8 L 303.2,452.3 L 307.0,451.0 L 310.8,454.1 L 315.4,451.8 L 318.1,452.8 L 335.9,441.0 L 343.6,439.3 L 351.1,435.0 L 357.6,424.9 L 361.4,423.8 L 370.1,414.1 L 374.7,406.5 L 377.8,395.8 L 384.5,387.0 L 383.4,381.5 L 385.7,380.3 L 387.2,374.9 L 386.6,368.0 L 388.4,367.3 L 391.9,359.1 L 394.8,357.3 L 392.1,339.8 L 396.0,336.9 L 389.0,336.0 L 380.2,318.9 L 373.1,311.5 L 380.8,292.4 L 393.2,283.3 L 392.2,276.2 L 393.9,274.5 L 391.5,273.8 L 392.3,272.2 L 386.3,262.6 L 366.9,241.7 L 366.4,238.0 L 373.4,233.9 L 375.6,225.9 L 372.3,219.7 L 359.7,205.5 L 345.7,200.9 L 336.3,204.5 L 322.3,204.4 L 318.7,206.3 L 317.7,210.4 L 300.8,209.2 L 297.6,210.5 L 297.6,214.1 L 280.2,218.3 L 275.8,223.5 L 277.1,226.2 L 269.6,231.6 L 268.7,236.6 L 273.4,241.5 L 272.5,245.2 L 271.0,246.6 L 264.3,246.2 L 263.3,249.9 L 260.7,251.3 L 256.6,248.2 L 258.0,243.3 L 248.3,237.2 L 239.4,234.9 L 236.8,235.2 L 233.5,240.0 L 227.9,243.5 L 221.8,241.5 L 215.5,245.7 L 212.4,243.0 L 207.0,242.4 L 202.9,249.8 L 195.8,247.6 L 192.9,244.1 L 197.4,241.4 L 197.5,237.7 L 205.3,230.2 L 204.6,224.2 L 213.0,222.0 L 214.8,216.7 L 220.9,214.7 L 220.9,211.9 L 226.1,205.4 L 230.9,205.3 L 236.0,198.9 L 242.8,197.3 L 248.5,193.5 L 249.5,190.2 L 251.8,193.8 L 262.8,192.9 L 273.7,183.8 L 276.6,183.5 L 279.4,178.3 L 286.1,172.4 L 293.6,172.2 L 309.4,159.8 L 310.4,155.8 L 313.7,153.3 L 312.7,150.1 L 314.3,147.5 L 314.2,140.8 L 313.3,143.7 L 310.4,133.4 L 312.5,129.1 L 318.3,128.2 L 331.2,121.0 L 340.7,110.7 L 350.5,105.5 L 350.8,99.1 L 354.6,95.7 L 364.5,93.1 L 384.1,79.4 L 386.0,80.8 L 391.2,77.1 L 392.5,78.7 L 391.4,76.6 L 395.1,79.7 L 391.3,75.2 L 391.7,71.4 L 393.4,72.3 L 392.1,73.8 L 397.3,73.8 L 396.3,75.8 L 408.1,75.7 L 409.9,77.6 L 407.6,78.5 L 410.1,79.7 L 412.9,77.6 L 431.3,74.0 L 442.5,67.8 L 448.0,67.7 L 450.5,64.3 L 453.2,64.2 L 456.1,67.3 L 454.6,67.3 L 460.2,70.4 L 453.3,62.8 L 445.7,64.0 L 449.1,60.4 L 452.4,62.0 L 448.2,59.8 L 449.6,58.8 L 448.2,54.3 L 444.1,52.4 L 444.0,50.4 L 447.1,49.1 L 457.3,50.8 L 461.3,55.9 L 472.7,53.9 L 477.4,57.2 L 484.8,58.8 L 485.8,63.0 L 492.1,65.8 L 520.9,67.2 L 540.9,60.3 L 552.7,52.2 L 564.5,49.3 L 565.8,51.5 L 592.0,43.2 L 605.2,42.9 L 612.4,47.0 L 608.0,46.8 L 604.4,49.3 L 608.0,46.8 L 612.5,47.0 L 614.7,49.5 L 613.2,50.9 L 619.0,55.2 L 619.3,57.8 L 618.8,55.7 L 620.8,56.9 L 620.6,49.3 L 628.0,52.7 L 631.4,50.7 L 645.9,60.4 L 667.9,79.5 L 685.6,82.9 L 707.1,96.2 L 714.7,104.0 L 715.1,108.4 L 708.6,112.6 L 694.2,116.6 L 697.1,118.9 L 698.6,117.9 L 703.0,124.2 L 708.7,128.2 L 720.1,131.7 L 726.3,128.2 L 725.6,131.9 L 723.3,132.7 L 723.7,136.1 L 720.4,138.7 L 727.9,142.5 L 726.2,145.1 L 731.9,150.0 L 728.8,150.7 L 732.6,152.3 L 729.5,155.5 L 733.6,158.4 L 732.1,163.6 L 736.9,165.4 L 739.1,172.8 L 741.9,174.0 L 749.9,175.5 L 754.7,173.1 L 759.0,173.8 L 760.7,176.6 L 769.8,177.5 L 767.6,182.5 L 770.8,187.2 L 770.1,197.5 L 761.9,198.7 L 749.2,209.2 L 745.7,202.3 L 738.1,195.5 L 737.7,192.3 L 725.7,190.8 L 724.6,188.7 L 712.9,193.7 L 704.1,189.2 L 699.8,182.8 L 697.8,186.3 L 690.0,187.4 L 685.9,194.5 L 682.3,197.3 L 671.2,198.4 L 665.6,196.7 L 661.7,201.0 L 660.5,209.7 L 651.3,210.6 L 654.4,212.9 L 651.2,213.0 L 649.7,215.1 L 656.6,219.2 L 646.6,226.3 L 650.0,228.1 L 645.3,236.4 L 645.2,240.7 L 649.4,241.0 L 647.0,243.5 L 647.9,245.6 L 659.5,249.2 L 662.7,245.8 L 668.2,244.9 L 666.8,248.1 L 670.3,253.5 L 679.2,253.1 L 677.9,254.5 L 681.7,253.9 L 683.0,256.2 L 687.9,254.9 L 689.7,258.1 L 686.6,260.5 L 694.9,263.4 L 693.5,265.4 L 695.3,267.0 L 702.6,267.6 L 703.8,273.2 L 698.9,273.6 L 699.1,269.6 L 692.6,271.9 L 686.7,270.4 L 686.4,276.7 L 683.1,281.4 L 679.5,282.2 L 679.9,287.9 L 681.9,287.9 L 678.9,292.6 L 680.1,294.9 L 678.3,297.9 L 669.1,300.1 L 675.9,303.9 L 672.1,304.5 L 666.3,311.5 L 667.5,315.0 L 659.5,320.0 L 653.3,320.0 L 654.6,320.7 L 653.8,324.2 L 661.2,335.8 L 654.0,338.4 L 655.1,340.8 L 668.6,338.3 L 669.9,336.6 L 666.2,334.6 L 668.1,332.3 L 675.7,330.9 L 682.4,325.0 L 688.7,323.6 L 692.6,326.3 L 692.8,331.1 L 694.4,331.7 L 691.7,343.6 L 683.6,342.2 L 682.5,346.5 L 684.5,350.2 L 661.2,351.1 L 641.7,360.0 L 637.0,359.2 L 627.2,367.4 L 626.2,371.2 L 628.2,372.3 L 627.5,375.5 L 630.6,378.8 L 642.1,383.0 L 647.3,386.5 L 647.1,389.2 L 649.0,390.8 L 661.5,394.6 L 675.6,404.9 L 681.4,406.3 L 700.0,402.6 L 705.6,404.6 L 702.1,411.7 L 692.5,414.4 L 692.4,422.2 L 689.2,429.6 L 707.0,433.1 L 702.6,433.6 L 702.7,435.7 L 699.2,436.6 L 691.2,435.0 L 682.6,437.7 L 679.1,440.0 L 680.3,444.4 L 672.1,445.2 L 672.9,447.5 L 670.7,452.2 L 672.4,453.3 L 664.0,457.0 L 666.0,463.8 L 658.1,471.1 L 664.8,474.2 L 669.4,471.2 L 672.2,473.8 L 666.3,477.9 L 654.3,480.2 L 652.6,483.8 L 663.1,487.8 L 662.9,491.8 L 658.1,492.6 L 659.2,497.8 L 657.1,501.5 L 659.8,506.0 L 667.1,511.5 L 668.4,515.2 L 664.7,521.4 L 668.2,524.6 L 673.4,535.3 L 685.4,544.8 L 690.4,551.7 L 690.5,555.6 L 693.0,556.9 L 699.4,555.5 L 700.7,563.5 L 706.2,563.1 L 723.2,556.2 L 724.3,553.3 L 727.3,553.2 L 729.5,555.7 L 736.6,557.4 L 731.8,560.3 L 737.5,559.0 L 739.1,563.2 L 743.4,563.4 L 748.9,568.8 L 758.3,572.3 L 755.3,577.5 L 766.6,588.0 L 772.0,583.6 L 773.3,586.7 L 778.3,585.1 L 779.5,589.0 L 782.4,587.3 L 787.2,588.3 L 791.7,591.7 L 788.7,593.0 L 789.0,597.3 L 783.7,600.1 L 786.2,603.1 L 783.1,612.2 L 788.1,617.6 L 785.4,623.9 L 781.1,626.0 L 786.5,634.8 L 785.8,636.6 L 782.2,636.7 L 786.3,639.3 L 785.4,643.7 L 789.8,646.0 L 782.3,648.1 L 781.4,650.1 L 783.7,651.3 L 787.2,649.6 L 783.1,655.6 L 786.7,656.7 L 786.5,660.6 L 789.8,664.5 L 787.2,674.8 L 777.5,678.8 L 774.6,683.5 L 770.9,684.0 L 770.9,687.1 L 765.1,696.6 L 764.8,694.5 L 758.3,699.3 L 755.6,699.3 L 753.8,695.5 L 755.1,693.3 L 753.9,695.8 L 752.3,695.3 L 755.1,698.3 L 752.9,695.8 L 753.7,699.8 L 750.9,699.1 L 750.8,696.3 L 747.8,698.1 L 747.1,702.3 L 745.6,701.6 L 747.3,698.4 L 745.4,701.5 L 746.2,698.9 L 743.8,700.1 L 743.0,698.3 L 749.3,694.4 L 728.2,703.2 L 733.4,702.4 L 733.8,704.3 L 724.3,703.7 L 722.3,702.3 L 723.3,701.5 L 728.7,701.4 L 732.1,698.3 L 738.7,696.7 L 742.4,691.2 L 735.5,694.2 L 737.7,695.3 L 731.3,696.9 L 729.1,700.0 L 725.7,699.7 L 727.7,698.5 L 727.2,696.2 L 735.2,693.2 L 728.2,695.0 L 723.5,698.7 L 722.8,697.0 L 725.0,697.8 L 725.1,695.9 L 720.7,698.3 L 710.6,697.0 L 698.1,698.8 L 676.1,707.5 L 653.5,722.3 L 653.0,725.7 L 646.9,729.5 L 644.9,735.1 L 645.6,742.5 L 637.0,744.7 L 637.6,746.2 L 629.2,743.7 L 614.0,749.6 L 605.4,747.2 L 596.6,751.2 L 589.4,750.2 L 577.5,751.8 L 573.4,749.1 L 534.6,744.3 L 525.0,732.0 L 513.0,723.8 L 507.5,717.1 L 503.3,716.5 L 500.3,718.5 L 489.6,716.6 L 478.2,704.8 L 474.8,695.1 L 477.1,694.8 L 474.7,694.9 L 471.2,687.8 L 472.2,687.1 L 461.7,682.0 L 463.6,678.6 L 451.0,670.2 L 446.7,671.8 L 439.8,668.6 L 430.7,670.4 L 423.0,669.6 L 418.7,672.2 L 414.3,679.3 L 413.9,682.4 L 418.7,685.0 L 403.1,686.0 L 399.4,684.2 L 391.8,688.1 L 374.4,684.1 L 366.5,687.5 L 368.2,693.4 L 354.5,691.7 L 352.1,693.0 L 352.1,695.9 L 343.2,693.7 L 328.8,686.4 L 323.1,687.8 L 315.8,685.6 L 322.0,686.3 L 326.2,682.1 L 325.5,676.1 L 320.7,669.1 L 335.3,661.6 L 338.0,656.7 L 337.5,653.5 L 344.6,650.5 L 339.2,649.8 L 340.1,647.6 L 351.2,644.8 L 347.5,643.5 L 340.7,644.8 L 328.8,651.4 L 324.8,649.7 L 326.3,649.3 L 319.1,648.9 L 303.4,637.7 L 301.0,630.8 L 302.5,637.6 L 300.1,635.6 L 299.8,637.2 L 299.2,626.5 L 299.0,637.1 L 297.6,633.3 L 294.9,632.3 L 296.6,626.9 L 294.2,631.4 L 297.4,635.5 L 293.7,633.9 L 292.9,630.1 L 292.2,633.9 L 285.8,634.1 L 286.5,629.1 L 297.8,625.1 L 295.8,624.1 L 289.5,626.5 L 293.4,623.7 L 294.2,620.2 L 281.4,627.6 L 280.1,626.5 L 281.1,627.8 L 274.3,628.3 L 250.7,623.4 L 220.9,627.0 L 214.2,632.5 L 215.7,637.0 L 218.7,637.9 L 213.0,642.6 L 212.2,645.7 L 213.9,647.4 Z"
)

WALES_ANGLESEY_PATH = (
    "M 266.4,94.3 L 259.7,91.0 L 260.3,88.1 L 256.2,89.7 L 257.2,86.5 L 253.5,83.8 L 245.7,85.9 L 243.4,84.4 L 242.0,86.1 L 237.5,84.3 L 236.2,82.9 L 237.6,80.0 L 233.0,75.4 L 235.1,72.1 L 225.6,69.2 L 216.8,70.9 L 215.3,68.6 L 218.0,64.8 L 211.9,62.1 L 216.8,59.7 L 217.4,55.9 L 222.7,57.6 L 234.8,53.3 L 227.0,57.0 L 232.6,59.1 L 233.2,57.4 L 234.7,58.2 L 231.5,61.6 L 233.6,59.8 L 238.0,62.6 L 247.2,63.6 L 249.7,62.2 L 249.4,59.4 L 244.7,55.9 L 244.3,53.4 L 248.5,51.0 L 249.7,46.0 L 248.1,45.5 L 252.5,38.1 L 247.2,28.8 L 250.0,26.1 L 260.2,26.1 L 260.4,24.4 L 265.4,22.4 L 265.2,24.6 L 267.1,25.3 L 269.2,23.0 L 271.9,24.1 L 274.7,20.5 L 277.9,23.3 L 282.6,23.6 L 282.1,21.6 L 283.8,21.0 L 282.5,20.4 L 289.6,17.8 L 295.5,18.8 L 296.3,21.0 L 298.1,19.3 L 306.1,20.0 L 309.3,23.4 L 314.1,21.5 L 315.7,23.3 L 318.1,22.1 L 327.2,24.5 L 328.6,22.4 L 334.2,31.8 L 333.5,37.6 L 337.8,42.3 L 345.1,43.6 L 343.4,45.9 L 344.9,49.9 L 350.6,58.0 L 364.0,61.4 L 369.6,61.5 L 376.2,57.3 L 398.8,60.5 L 381.4,81.4 L 371.1,85.7 L 363.8,92.8 L 360.3,91.8 L 352.0,95.5 L 349.5,98.6 L 348.3,105.4 L 340.8,107.7 L 320.9,122.1 L 317.4,122.7 L 317.0,126.6 L 302.3,123.8 L 299.7,122.2 L 303.9,122.5 L 298.1,120.1 L 295.0,120.5 L 292.5,124.1 L 294.5,120.2 L 291.0,114.7 L 284.5,116.3 L 284.1,114.0 L 281.0,113.7 L 282.3,112.0 L 278.7,107.6 L 271.6,108.7 L 272.2,105.5 L 266.2,104.6 L 268.9,97.9 L 267.3,97.9 L 266.4,94.3 Z"
)

# Region zones are simple rectangular bands. We apply a clipPath of the
# real coastline to the whole zone group at render time, so each band
# only paints inside Wales — no need to redraw 50+ coastline points per
# zone. Inland boundaries between adjacent regions are approximate.
#
# Layering: the first five zones (Mid Wales, Pembs, Carms, BB, Wye) form
# a complete tiling of south Wales with no gaps. Gower and Valleys & Vale
# are then drawn on TOP as opaque overlays — since `.wales-zone` fills
# with parchment under the colour-mix, the overlays cleanly cover the
# bands beneath. The English Borders zone is rendered OUTSIDE the clip
# (it deliberately extends east of the Wales coastline).
#
# Centroids (cx, cy) are placed well inland so labels stay clear of the
# clipped coast — Mid Wales ≈ Powys plateau (not Aberystwyth coast);
# Pembrokeshire ≈ Haverfordwest; Carmarthenshire ≈ Tywi valley.
WALES_MAP_ZONES = [
    {
        "full": "Mid Wales (Powys & Ceredigion)",
        # Full-width band, clipped to Wales between y=270 (≈ Aberystwyth)
        # and y=478 (≈ northern edge of Pembs/Brecon Beacons).
        "d":    "M -20,270 L 840,270 L 840,478 L -20,478 Z",
        "cx":   470, "cy": 380,
    },
    {
        "full": "Pembrokeshire (South)",
        # SW peninsula. East boundary at x=235 (Carmarthen Bay shoulder).
        "d":    "M -20,478 L 235,478 L 235,790 L -20,790 Z",
        "cx":   135, "cy": 615,
        "label": ["South", "Pembrokeshire"],
    },
    {
        "full": "Carmarthenshire & West Wales",
        # Full vertical strip between Pembs and Brecons. The Gower
        # overlay (drawn later) covers the SW corner; the label sits
        # clear of that overlay in the upper Tywi valley.
        "d":    "M 235,478 L 460,478 L 460,790 L 235,790 Z",
        "cx":   350, "cy": 560,
    },
    {
        "full": "Brecon Beacons / Bannau Brycheiniog",
        # Inland uplands. No coast — fully interior. Valleys & Vale
        # overlays the southern strip of this rect.
        "d":    "M 460,478 L 660,478 L 660,640 L 460,640 Z",
        "cx":   560, "cy": 555,
    },
    {
        "full": "Wye Valley & Monmouthshire",
        # East border, inside the Welsh side. Goes all the way south.
        "d":    "M 660,478 L 800,478 L 800,790 L 660,790 Z",
        "cx":   720, "cy": 600,
        "label": ["Wye Valley"],
    },
    # ── Overlays drawn on top of the bands above ──────────────────────
    {
        "full": "Gower & Swansea Bay",
        # Peninsula band hanging off the south of Carmarthenshire.
        # Drawn AFTER Carmarthenshire so it cleanly overlays the SW.
        "d":    "M 320,655 L 460,655 L 460,790 L 250,790 Z",
        "cx":   355, "cy": 720,
        "label": ["Gower &", "Swansea"],
    },
    {
        "full": "Valleys & Vale of Glamorgan",
        # SE strip below Brecons, east of Gower. Extended west to x=460
        # to eliminate the polygon gap with Brecon Beacons / Carms.
        "d":    "M 460,640 L 660,640 L 660,790 L 460,790 Z",
        "cx":   560, "cy": 715,
        "label": ["Valleys &", "Vale"],
    },
    {
        "full": "English Borders (Forest of Dean & Herefordshire)",
        # Drawn OUTSIDE the wales clip — extends beyond the border by design.
        "d":    "M 735,485 L 780,475 L 815,505 L 815,575 L 780,615 "
                "L 735,618 L 725,580 L 735,525 Z",
        "cx":   775, "cy": 548,
        "label": ["Borders"],
    },
]
def _zone_shape_svg(z: dict, counts: dict, *, is_border: bool = False) -> list[str]:
    """Render the clickable path for one region (no labels).

    Sits INSIDE the wales-clip group so the band is silhouetted by the
    real coastline. Labels live in a separate group rendered outside the
    clip so they never get clipped.
    """
    meta   = REGION_META.get(z["full"], {})
    accent = meta.get("accent", "#4a6b3e")
    short  = meta.get("short", z["full"])
    count  = counts.get(z["full"], 0)
    extra_cls = " zone-over-border" if is_border else ""
    return [
        f'<g class="wales-zone-group wales-zone-shape" '
        f'data-region-full="{esc(z["full"])}" '
        f'style="--zone-color: {accent}">',
        f'<path d="{z["d"]}" class="wales-zone{extra_cls}" '
        f'role="button" tabindex="0" '
        f'aria-label="{esc(short)} — {count} walks"></path>',
        '</g>',
    ]


def _zone_labels_svg(z: dict, counts: dict) -> list[str]:
    """Render the label + count text for one region (no path).

    Drawn OUTSIDE the wales-clip group so the text never gets clipped by
    the coastline — important for centroids that sit near the coast or
    on the western edge (e.g. Mid Wales centroid above Aberystwyth).
    """
    meta   = REGION_META.get(z["full"], {})
    accent = meta.get("accent", "#4a6b3e")
    short  = meta.get("short", z["full"])
    count  = counts.get(z["full"], 0)
    label_lines = z.get("label") or [short]
    line_h = 15
    parts = [
        f'<g class="wales-zone-group wales-zone-labels" '
        f'data-region-full="{esc(z["full"])}" '
        f'style="--zone-color: {accent}">',
        f'<text x="{z["cx"]}" y="{z["cy"]}" class="zone-label" '
        f'text-anchor="middle" pointer-events="none">',
    ]
    for i, line in enumerate(label_lines):
        if i == 0:
            parts.append(f'<tspan x="{z["cx"]}">{esc(line)}</tspan>')
        else:
            parts.append(f'<tspan x="{z["cx"]}" dy="{line_h}">{esc(line)}</tspan>')
    parts.append('</text>')
    count_y = z["cy"] + line_h * (len(label_lines) - 1) + 16
    parts.append(
        f'<text x="{z["cx"]}" y="{count_y}" class="zone-count" '
        f'text-anchor="middle" pointer-events="none">{count} walks</text>'
    )
    parts.append('</g>')
    return parts


def wales_map_svg(counts: dict) -> str:
    """Render the interactive Wales map as inline SVG.

    Layering (bottom -> top):
      1. Sea/background rect
      2. <g clip-path="url(#wales-clip)">  ← clipped to real coastline
         a. Parchment land base
         b. North Wales "coming soon" band
         c. Zone SHAPES (paths only) — base bands then opaque overlays
      3. English Borders zone shape (outside clip — extends past Wales)
      4. Coastline strokes (mainland + Anglesey)
      5. Zone LABELS + counts (OUTSIDE clip, so they're never clipped)
      6. North Wales "coming soon" label
    """
    parts = [
        f'<svg viewBox="{WALES_MAP_VIEWBOX}" class="wales-svg" '
        f'role="img" aria-label="Map of Wales — click a region to filter walks" '
        f'preserveAspectRatio="xMidYMid meet">',
        # Defs: clipPath using the actual ONS Wales coastline. Both shapes are
        # included so paths drawn inside the clip group also paint Anglesey.
        '<defs>',
        '<clipPath id="wales-clip">',
        f'<path d="{WALES_MAINLAND_PATH}"/>',
        f'<path d="{WALES_ANGLESEY_PATH}"/>',
        '</clipPath>',
        '</defs>',
        # 1. Sea / outer canvas — kept very subtle so the land pops.
        '<rect class="wales-sea" x="0" y="0" width="820" height="780"/>',
        # 2. Everything that should stop at the coastline.
        '<g clip-path="url(#wales-clip)">',
        # 2a. Parchment land fill spanning the full canvas; clip handles the rest.
        '<rect class="wales-land" x="-20" y="-20" width="860" height="820"/>',
        # 2b. North Wales "coming soon" band — full-width rect, clipped, so it
        # naturally takes the shape of the actual north Wales coastline + Anglesey.
        '<rect x="-20" y="-20" width="860" height="290" class="zone-future" '
        'aria-hidden="true"/>',
    ]
    # 2c. Zone shapes (paths only). Order in WALES_MAP_ZONES is significant:
    # base bands first (Mid Wales, Pembs, Carms, BB, Wye), then overlays
    # (Gower, Valleys & Vale) which paint on top with opaque fill.
    for z in WALES_MAP_ZONES:
        if z["full"].startswith("English Borders"):
            continue
        parts.extend(_zone_shape_svg(z, counts, is_border=False))
    parts.append('</g>')  # close clip group

    # 3. English Borders zone — shape outside the clip (extends past border).
    for z in WALES_MAP_ZONES:
        if not z["full"].startswith("English Borders"):
            continue
        parts.extend(_zone_shape_svg(z, counts, is_border=True))

    # 4. Coastline strokes — drawn on top of the regions, no fill, just outline.
    parts.append(
        f'<path d="{WALES_MAINLAND_PATH}" class="wales-coast" '
        f'pointer-events="none" aria-hidden="true"/>'
    )
    parts.append(
        f'<path d="{WALES_ANGLESEY_PATH}" class="wales-coast" '
        f'pointer-events="none" aria-hidden="true"/>'
    )

    # 5. Zone LABELS — outside the clip so they never get truncated by the
    # coastline. Order doesn't matter for labels (they don't overlap).
    for z in WALES_MAP_ZONES:
        parts.extend(_zone_labels_svg(z, counts))

    # 6. North Wales label, positioned over the visible north Wales body.
    parts.append(
        '<text x="430" y="145" class="future-label" text-anchor="middle" '
        'pointer-events="none">North Wales</text>'
    )
    parts.append(
        '<text x="430" y="170" class="future-sub" text-anchor="middle" '
        'pointer-events="none">coming soon</text>'
    )

    parts.append('</svg>')
    return "\n".join(parts)

# ---------------------------------------------------------------------------
# Photo bank: per-feature Unsplash image IDs for each walk's detail gallery.
# Any bad IDs fall back to a plain background via the img onerror handler.
# ---------------------------------------------------------------------------
PHOTO_BANK = {
    "Mountain / Summit":     ["photo-1464822759023-fed622ff2c3b", "photo-1506905925346-21bda4d32df4", "photo-1519681393784-d120267933ba"],
    "Waterfall":             ["photo-1501286353178-1ec881214838", "photo-1432889490240-84df33d47091", "photo-1469474968028-56623f02e42e"],
    "Beach / Coastal":       ["photo-1507525428034-b723cf961d3e", "photo-1476514525535-07fb3b4ae5f1", "photo-1439405326854-014607f694d7"],
    "Lake / Tarn / Pond":    ["photo-1506905925346-21bda4d32df4", "photo-1469474968028-56623f02e42e", "photo-1470252649378-9c29740c9fa8"],
    "Woodland / Forest":     ["photo-1441974231531-c6227db76b6e", "photo-1448375240586-882707db888b", "photo-1426604966848-d7adac402bff"],
    "River / Gorge":         ["photo-1469474968028-56623f02e42e", "photo-1500382017468-9049fed747ef", "photo-1441974231531-c6227db76b6e"],
    "Castle":                ["photo-1548248823-ce16a73b6d49", "photo-1583001931096-959e9a1a6223", "photo-1578668582937-3e0b64c1aec7"],
    "Abbey / Church":        ["photo-1548248823-ce16a73b6d49", "photo-1583001931096-959e9a1a6223", "photo-1509248961158-e54f6934749c"],
    "Prehistoric / Roman":   ["photo-1583001931096-959e9a1a6223", "photo-1548248823-ce16a73b6d49", "photo-1447069387593-a5de0862481e"],
    "Industrial Heritage":   ["photo-1583001931096-959e9a1a6223", "photo-1469854523086-cc02fe5d8800", "photo-1441974231531-c6227db76b6e"],
    "Wildlife / Nature Reserve": ["photo-1441974231531-c6227db76b6e", "photo-1447069387593-a5de0862481e", "photo-1448375240586-882707db888b"],
    "Waymarked Long-Distance Path": ["photo-1441974231531-c6227db76b6e", "photo-1464822759023-fed622ff2c3b", "photo-1519681393784-d120267933ba"],
}
NATURE_POOL = [
    "photo-1464822759023-fed622ff2c3b", "photo-1506905925346-21bda4d32df4",
    "photo-1441974231531-c6227db76b6e", "photo-1501286353178-1ec881214838",
    "photo-1507525428034-b723cf961d3e", "photo-1519681393784-d120267933ba",
    "photo-1469474968028-56623f02e42e", "photo-1470252649378-9c29740c9fa8",
    "photo-1448375240586-882707db888b", "photo-1426604966848-d7adac402bff",
    "photo-1476514525535-07fb3b4ae5f1", "photo-1470071459604-3b5ec3a7fe05",
]

def pick_images(tags, region, walk_id):
    """Deterministically pick 3 unique image URLs for a walk."""
    picked = []
    for tag in tags or []:
        for pid in PHOTO_BANK.get(tag, []):
            if pid not in picked:
                picked.append(pid)
                if len(picked) == 3:
                    break
        if len(picked) == 3:
            break
    # Top up with the region's hero photo if we still need more
    region_photo = REGION_META.get(region, {}).get("image", "")
    m = re.search(r"(photo-[\w-]+)", region_photo)
    if m and m.group(1) not in picked and len(picked) < 3:
        picked.append(m.group(1))
    # Finally, top up from a rotated nature pool for variety
    start = (abs(hash(str(walk_id))) % len(NATURE_POOL)) if walk_id is not None else 0
    safety = 0
    while len(picked) < 3 and safety < len(NATURE_POOL) * 2:
        cand = NATURE_POOL[(start + safety) % len(NATURE_POOL)]
        if cand not in picked:
            picked.append(cand)
        safety += 1
    return [u(pid, w=800, q=70) for pid in picked[:3]]


# Populate each walk's gallery from the photo cache. Walks without cached
# photos get an EMPTY gallery rather than a random Unsplash fallback — stock
# photos keyed on tags produced jarring mismatches ("Abbey / Church" → clown
# masks, pyramids, honeycomb, etc.) because Unsplash occasionally recycles
# photo IDs. "No photos" is a far better failure mode than "wrong photos".
#
# To populate galleries, run the GitHub Actions workflow — fetch_photos.py
# pulls geo-searched photos from Wikimedia Commons (free, CC-BY-SA / CC-BY)
# using each walk's postcode.
photo_used = 0
for rec in data:
    cached = _photo_cache.get(str(rec.get("id")), {}).get("photos") or []
    if cached:
        # Prefer the server-scaled `thumb` over the full-size `url` —
        # Commons originals can be 20MB+ JPEGs that would crush page load.
        rec["images"] = [p.get("thumb") or p["url"] for p in cached[:3]]
        rec["photo_credits"] = [
            {
                "photographer": p.get("photographer", ""),
                "page_url":     p.get("page_url", ""),
                "title":        p.get("title", ""),
                "license":      p.get("license", "CC BY-SA"),
                "license_url":  p.get("license_url", "https://creativecommons.org/licenses/by-sa/4.0/"),
                "source":       p.get("source", "Wikimedia Commons"),
            }
            for p in cached[:3]
        ]
        photo_used += 1
    else:
        # Intentionally empty — no stock-photo fallback.
        rec["images"] = []
        rec["photo_credits"] = []
print(f"  using cached photos: {photo_used}/{len(data)}; no photos: {len(data) - photo_used}")


# ---------------------------------------------------------------------------
# Curated featured walks (one each: summit, coast, waterfall, ruin)
# ---------------------------------------------------------------------------
FEATURED_PICKS = [
    {
        "name": "Pen y Fan Circular (Motorway Route)",
        "image": u("photo-1464822759023-fed622ff2c3b", w=1400),
        "kicker": "The summit",
    },
    {
        "name": "Rhossili Down & Worm's Head",
        "image": u("photo-1507525428034-b723cf961d3e", w=1400),
        "kicker": "The coast",
    },
    {
        "name": "Four Waterfalls Walk (Sgwd yr Eira)",
        "image": u("photo-1519681393784-d120267933ba", w=1400),
        "kicker": "The waterfall",
    },
    {
        "name": "Tintern Abbey & Devil's Pulpit",
        "image": u("photo-1548248823-ce16a73b6d49", w=1400),
        "kicker": "The ruin",
    },
]

def _find_by_name(name):
    for w in data:
        if w["name"] == name:
            return w
    return None

featured_walks = []
for pick in FEATURED_PICKS:
    w = _find_by_name(pick["name"])
    if w:
        featured_walks.append({**w, "image": pick["image"], "kicker": pick["kicker"]})

# Region walk-count map for the region tiles
from collections import Counter
region_counts = Counter(w["region"] for w in data)
region_tiles = []
for region, meta in REGION_META.items():
    if region_counts.get(region, 0) > 0:
        region_tiles.append({
            "region": region,
            "short": meta["short"],
            "tagline": meta["tagline"],
            "image": meta["image"],
            "accent": meta["accent"],
            "count": region_counts[region],
        })

# Filter-panel values
regions = sorted({w["region"] for w in data if w["region"]})
difficulties = ["Easy", "Easy/Moderate", "Moderate", "Moderate/Hard", "Hard", "Very Hard"]
all_tags = sorted({t for w in data for t in w["tags"]})
max_miles = max(w["miles"] for w in data) if data else 10
max_drive = max(w["drive"] for w in data if isinstance(w["drive"], (int, float))) or 180
near_count = sum(1 for w in data if isinstance(w["drive"], int) and w["drive"] <= 60)

# Accent colours keyed by full region name, for embedding in JSON for JS
region_accent_js = {r: REGION_META.get(r, {}).get("accent", "#4a6b3e") for r in regions}
region_short_js = {r: REGION_META.get(r, {}).get("short", r) for r in regions}

# ---------------------------------------------------------------------------
# Server-side HTML snippets for sections that don't need to re-render
# (featured cards and region tiles are rendered once in Python).
# ---------------------------------------------------------------------------
def esc(s):
    return html.escape(str(s if s is not None else ""), quote=True)

def featured_card(w):
    accent = REGION_META.get(w["region"], {}).get("accent", "#4a6b3e")
    short_reg = REGION_META.get(w["region"], {}).get("short", w["region"])
    return f'''
    <button class="feat" type="button" data-walk-name="{esc(w["name"])}" style="--img:url('{w["image"]}');--accent:{accent}" aria-label="Open walk: {esc(w["name"])}">
      <span class="region-chip">{esc(w["kicker"])} &middot; {esc(short_reg)}</span>
      <h3>{esc(w["name"])}</h3>
      <div class="f-stats">
        <span><b>{w["miles"]}</b>mi</span>
        <span><b>{w["elev"]}</b>m</span>
        <span><b>{w["time"]}</b>hrs</span>
        <span><b>{w["difficulty"]}</b></span>
      </div>
      <span class="feat-cta">View walk →</span>
    </button>'''

def region_tile(t):
    return f'''
    <button class="region-tile" type="button" data-region-short="{esc(t["short"])}"
      style="--img:url('{t["image"]}');--accent:{t["accent"]}">
      <div>
        <h4>{esc(t["short"])}</h4>
        <div class="r-tagline">{esc(t["tagline"])}</div>
      </div>
      <span class="r-count">{t["count"]} <small style="font-family:Inter,sans-serif;font-size:.7rem;letter-spacing:.16em;text-transform:uppercase;color:rgba(250,246,238,.75);font-weight:600">walks</small></span>
    </button>'''

featured_html = "\n".join(featured_card(w) for w in featured_walks)
region_tiles_html = "\n".join(region_tile(t) for t in region_tiles)
wales_map_html = wales_map_svg(region_counts)

# ---------------------------------------------------------------------------
# HTML template
# ---------------------------------------------------------------------------
LOGO_SVG = '''<svg class="brand-mark" viewBox="0 0 48 40" xmlns="http://www.w3.org/2000/svg" aria-hidden="true">
  <path d="M2 34 L18 8 L24 15 L30 6 L46 34 Z" fill="#2d3a2a"/>
  <path d="M2 34 L12 20 L20 34 Z" fill="#4a6b3e"/>
  <path d="M4 31 Q 15 29 24 26 T 44 29" stroke="#d9a64e" stroke-width="1.8" fill="none" stroke-linecap="round"/>
  <circle cx="36" cy="10" r="3" fill="#d9a64e" opacity="0.92"/>
</svg>'''

LOGO_SVG_FOOTER = '''<svg width="36" height="30" viewBox="0 0 48 40" xmlns="http://www.w3.org/2000/svg" aria-hidden="true">
  <path d="M2 34 L18 8 L24 15 L30 6 L46 34 Z" fill="currentColor" opacity=".85"/>
  <path d="M4 31 Q 15 29 24 26 T 44 29" stroke="#d9a64e" stroke-width="1.8" fill="none" stroke-linecap="round"/>
  <circle cx="36" cy="10" r="3" fill="#d9a64e"/>
</svg>'''

HTML = r"""<!DOCTYPE html>
<html lang="en-GB">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<meta name="theme-color" content="#2d3a2a">
<title>South Wales Walks — 167 curated routes across Wales &amp; the Borders</title>
<meta name="description" content="A hand-compiled guide to 167 walks across the Brecon Beacons, Gower, Pembrokeshire, Wye Valley, Mid Wales, Carmarthenshire and the Forest of Dean. Filter by distance, difficulty, pubs, waterfalls, dogs and more.">
<meta property="og:title" content="South Wales Walks — 167 curated routes">
<meta property="og:description" content="Hand-compiled walks across the Brecon Beacons, Gower, Pembrokeshire, Wye Valley, Mid Wales, and the Forest of Dean.">
<meta property="og:type" content="website">
<meta name="twitter:card" content="summary_large_image">

<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Fraunces:ital,opsz,wght@0,9..144,400;0,9..144,500;0,9..144,600;1,9..144,400;1,9..144,500&family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">

<style>
:root{
  --bg:#f4ecdd;
  --paper:#fbf6ec;
  --card:#ffffff;
  --ink:#1b1f1a;
  --ink-soft:#3a4236;
  --muted:#76745f;
  --stone:#d7cfc0;
  --border:#e2dbc9;
  --border-soft:#ebe4d2;
  --moss:#2f4530;
  --moss-2:#4a6b3e;
  --fern:#8ba67f;
  --bracken:#b45a2a;
  --bracken-dark:#8b3f17;
  --amber:#d9a64e;
  --slate:#263026;
  --cream:#faf6ee;
  --shadow-sm:0 1px 2px rgba(29,37,29,.06);
  --shadow-md:0 8px 24px rgba(29,37,29,.08);
  --shadow-lg:0 24px 60px rgba(29,37,29,.14);
}
*{box-sizing:border-box}
html{scroll-behavior:smooth}
body{margin:0;padding:0;background-color:var(--bg);color:var(--ink);
  background-image:url("data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' width='900' height='900' viewBox='0 0 900 900'><g fill='none' stroke='%233d5a3b' stroke-width='1' opacity='0.055'><path d='M0 180 Q225 140 450 180 T900 180'/><path d='M0 260 Q225 230 450 260 T900 260'/><path d='M0 340 Q225 320 450 340 T900 340'/><path d='M0 420 Q225 400 450 420 T900 420'/><path d='M0 500 Q225 490 450 500 T900 500'/><path d='M0 580 Q225 570 450 580 T900 580'/><path d='M0 660 Q225 640 450 660 T900 660'/><path d='M0 740 Q225 720 450 740 T900 740'/></g></svg>");
  background-size:900px 900px;background-attachment:fixed;background-repeat:repeat;
  font:15px/1.55 "Inter",-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;
  -webkit-font-smoothing:antialiased;text-rendering:optimizeLegibility;
}
h1,h2,h3,h4,h5{margin:0;font-family:"Fraunces",Georgia,serif;font-weight:500;letter-spacing:-.01em;line-height:1.1}
img{max-width:100%;display:block}
a{color:inherit;text-decoration:none}
button{font-family:inherit;color:inherit}

.container{max-width:1180px;margin:0 auto;padding:0 1.25rem}

/* ─── Navigation ─────────────────────────────────────────── */
nav.site{
  position:sticky;top:0;z-index:40;
  backdrop-filter:blur(12px);-webkit-backdrop-filter:blur(12px);
  background:rgba(244,236,221,.86);
  border-bottom:1px solid var(--border);
}
.nav-row{display:flex;align-items:center;justify-content:space-between;padding:.85rem 0;gap:.8rem}
.brand{display:flex;align-items:center;gap:.7rem}
.brand-mark{width:36px;height:32px;flex-shrink:0}
.brand-name{font-family:"Fraunces",serif;font-weight:500;font-size:1.02rem;letter-spacing:.01em;line-height:1}
.brand-name b{font-weight:600}
.brand-sub{display:block;font-family:"Inter",sans-serif;font-size:.62rem;letter-spacing:.25em;
  text-transform:uppercase;color:var(--muted);font-weight:600;margin-top:3px}
.nav-links{display:flex;gap:1.2rem;align-items:center;font-size:.88rem}
.nav-links a{color:var(--ink-soft);font-weight:500;transition:color .15s}
.nav-links a:hover{color:var(--bracken)}
.nav-cta{
  background:var(--slate);color:var(--cream)!important;padding:.55rem 1.05rem;border-radius:999px;
  font-size:.82rem;font-weight:600;border:1px solid var(--slate);transition:all .2s;
}
.nav-cta:hover{background:var(--bracken);border-color:var(--bracken)}
@media(max-width:720px){
  .nav-links a:not(.nav-cta){display:none}
  .brand-sub{display:none}
}

/* ─── Hero ───────────────────────────────────────────────── */
.hero{
  position:relative;min-height:min(78vh,680px);display:flex;align-items:flex-end;
  color:var(--cream);overflow:hidden;isolation:isolate;
  background:linear-gradient(135deg,#2d3a2a 0%,#3a4232 38%,#52654a 70%,#7a8661 100%);
}
.hero::before{
  content:"";position:absolute;inset:0;z-index:-2;
  background-image:url("https://images.unsplash.com/photo-1464822759023-fed622ff2c3b?auto=format&fit=crop&w=2200&q=70");
  background-size:cover;background-position:center 40%;opacity:.88;
}
.hero::after{
  content:"";position:absolute;inset:0;z-index:-1;
  background:linear-gradient(180deg,rgba(27,31,26,.28) 0%,rgba(27,31,26,.3) 40%,rgba(27,31,26,.82) 100%);
}
.hero-inner{width:100%;padding:6.5rem 0 3.2rem}
.hero-eyebrow{
  display:inline-flex;align-items:center;gap:.55rem;
  padding:.4rem .85rem;background:rgba(250,246,238,.15);
  backdrop-filter:blur(8px);border:1px solid rgba(250,246,238,.22);
  border-radius:999px;font-size:.72rem;letter-spacing:.2em;
  text-transform:uppercase;font-weight:700;margin-bottom:1.3rem;
}
.hero-eyebrow .dot{width:6px;height:6px;border-radius:50%;background:var(--amber)}
.hero h1{
  font-family:"Fraunces",serif;font-weight:400;
  font-size:clamp(2.4rem,5.2vw,4.6rem);line-height:1.02;
  max-width:19ch;font-variation-settings:"opsz" 140;
}
.hero h1 em{font-style:italic;color:var(--amber);font-weight:400;font-variation-settings:"opsz" 140}
.hero-lede{
  margin-top:1.25rem;max-width:54ch;font-size:1.08rem;line-height:1.55;
  color:rgba(250,246,238,.85);font-weight:400;
}
.hero-meta{margin-top:2rem;display:flex;gap:2.4rem;flex-wrap:wrap;
  font-size:.74rem;letter-spacing:.16em;text-transform:uppercase;
  color:rgba(250,246,238,.72);font-weight:600}
.hero-meta b{color:var(--amber);font-family:"Fraunces",serif;font-weight:500;
  font-size:1.5rem;margin-right:.45rem;letter-spacing:0;text-transform:none;display:inline-block}
.hero-cta{
  margin-top:2.4rem;display:inline-flex;align-items:center;gap:.6rem;
  padding:.8rem 1.4rem;background:var(--bracken);color:var(--cream)!important;
  border-radius:999px;font-size:.9rem;font-weight:600;letter-spacing:.02em;
  transition:all .2s;
}
.hero-cta:hover{background:var(--bracken-dark);transform:translateY(-1px);box-shadow:0 12px 28px rgba(180,90,42,.35)}

/* ─── Stats strip ────────────────────────────────────────── */
.stats{background:var(--paper);border-top:1px solid var(--border);border-bottom:1px solid var(--border)}
.stats-row{display:grid;grid-template-columns:repeat(4,1fr);gap:1rem;padding:1.6rem 0}
.stat{display:flex;flex-direction:column;gap:.2rem;padding:0 .2rem;border-left:2px solid var(--border-soft);padding-left:1rem}
.stat:first-child{border-left:0;padding-left:0}
.stat-n{font-family:"Fraunces",serif;font-weight:500;font-size:2rem;color:var(--moss);line-height:1;font-variation-settings:"opsz" 72}
.stat-l{font-size:.72rem;letter-spacing:.16em;text-transform:uppercase;color:var(--muted);font-weight:700;margin-top:.2rem}
@media(max-width:720px){
  .stats-row{grid-template-columns:repeat(2,1fr);gap:.8rem 1.2rem}
  .stat{border-left:0;padding-left:0}
  .stat:nth-child(3){border-top:1px solid var(--border-soft);padding-top:.8rem}
  .stat:nth-child(4){border-top:1px solid var(--border-soft);padding-top:.8rem}
}

/* ─── Sections ───────────────────────────────────────────── */
.section{padding:4rem 0}
@media(max-width:720px){.section{padding:3rem 0}}
.section.alt{background:var(--paper);border-top:1px solid var(--border);border-bottom:1px solid var(--border)}
.section-head{display:flex;justify-content:space-between;align-items:flex-end;gap:1.5rem;margin-bottom:2rem;flex-wrap:wrap}
.section-eyebrow{font-size:.7rem;letter-spacing:.24em;text-transform:uppercase;color:var(--bracken);font-weight:700;margin-bottom:.7rem}
.section-title{font-family:"Fraunces",serif;font-size:clamp(1.8rem,3.3vw,2.6rem);font-weight:500;max-width:24ch;line-height:1.08;font-variation-settings:"opsz" 96}
.section-title em{font-style:italic;color:var(--moss);font-weight:400}
.section-sub{font-size:.98rem;color:var(--muted);max-width:46ch;line-height:1.6}

/* ─── Featured walks ─────────────────────────────────────── */
.featured{display:grid;grid-template-columns:repeat(4,1fr);gap:1rem}
@media(max-width:1020px){.featured{grid-template-columns:repeat(2,1fr)}}
@media(max-width:520px){.featured{grid-template-columns:1fr}}
.feat{
  position:relative;border-radius:14px;overflow:hidden;min-height:340px;
  color:var(--cream);padding:1.4rem;
  display:flex;flex-direction:column;justify-content:flex-end;
  background:linear-gradient(135deg,var(--accent,#3d5a3b),#1b1f1a);
  transition:transform .35s cubic-bezier(.2,.7,.2,1),box-shadow .35s;
  box-shadow:var(--shadow-md);
  isolation:isolate;
  border:0;font:inherit;cursor:pointer;text-align:left;width:100%;
}
.feat:focus-visible{outline:3px solid var(--amber);outline-offset:3px}
.feat-cta{
  margin-top:1rem;display:inline-flex;align-items:center;gap:.4rem;
  font-size:.75rem;letter-spacing:.14em;text-transform:uppercase;font-weight:700;
  color:var(--amber);opacity:.85;transition:opacity .2s,transform .3s;
}
.feat:hover .feat-cta{opacity:1;transform:translateX(3px)}
.feat::before{
  content:"";position:absolute;inset:0;z-index:-2;
  background-image:var(--img);background-size:cover;background-position:center;
  opacity:.9;transition:transform .8s cubic-bezier(.2,.7,.2,1);
}
.feat::after{
  content:"";position:absolute;inset:0;z-index:-1;
  background:linear-gradient(180deg,rgba(27,31,26,.1) 0%,rgba(27,31,26,.82) 80%);
}
.feat:hover{transform:translateY(-4px);box-shadow:var(--shadow-lg)}
.feat:hover::before{transform:scale(1.06)}
.region-chip{
  display:inline-flex;align-items:center;padding:.3rem .7rem;
  background:rgba(250,246,238,.16);backdrop-filter:blur(8px);
  border:1px solid rgba(250,246,238,.2);
  border-radius:999px;font-size:.68rem;font-weight:700;letter-spacing:.14em;text-transform:uppercase;
  align-self:flex-start;margin-bottom:.9rem;
}
.feat h3{font-family:"Fraunces",serif;font-weight:500;font-size:1.35rem;line-height:1.12;max-width:18ch}
.f-stats{display:flex;gap:1.2rem;margin-top:1rem;font-size:.68rem;letter-spacing:.12em;
  text-transform:uppercase;color:rgba(250,246,238,.8);font-weight:700}
.f-stats b{color:var(--amber);font-family:"Fraunces",serif;font-weight:500;font-size:1.15rem;
  margin-right:.25rem;text-transform:none;letter-spacing:0;display:inline-block}

/* ─── Wales map ──────────────────────────────────────────── */
.wales-map-wrap{
  max-width:720px;margin:0 auto 2.2rem;padding:1rem 1.2rem 1.4rem;
  background:var(--card);border:1px solid var(--border);border-radius:16px;
  box-shadow:var(--shadow-sm);
}
.wales-svg{display:block;width:100%;height:auto;font-family:inherit}
/* Sea / land / coast — sets the cartographic backdrop. */
.wales-sea{fill:transparent}
.wales-land{fill:var(--bg)}
.wales-coast{
  fill:none;
  stroke:color-mix(in srgb, var(--ink-soft) 70%, transparent);
  stroke-width:1.1;stroke-linejoin:round;stroke-linecap:round;
}
.wales-zone-group{--zone-color:var(--moss)}
/* Fill blends the zone accent with the parchment background (opaque),
   so overlay zones (Gower, Valleys & Vale) cleanly cover the bands they
   sit on top of without showing two layers of tint stacking up. */
.wales-zone{
  fill:color-mix(in srgb, var(--zone-color) 24%, var(--bg));
  stroke:color-mix(in srgb, var(--zone-color) 55%, var(--bg));
  stroke-width:1;
  cursor:pointer;transition:fill .18s,stroke .18s,filter .18s;
}
.wales-zone:hover{fill:color-mix(in srgb, var(--zone-color) 44%, var(--bg));stroke:var(--zone-color);stroke-width:1.6}
.wales-zone:focus{outline:0;stroke:var(--zone-color);stroke-width:2.2;filter:drop-shadow(0 0 3px rgba(74,107,62,.35))}
.wales-zone-group.on .wales-zone{fill:color-mix(in srgb, var(--zone-color) 78%, var(--bg));stroke:var(--zone-color);stroke-width:1.8}
.wales-zone-group.on .zone-label,
.wales-zone-group.on .zone-count{fill:var(--cream)}
.zone-over-border{stroke-dasharray:5 4;opacity:.9}
/* paint-order:stroke gives the label a parchment halo so the coastline
   line doesn't cut visually through any letters that happen to sit near
   the coast. The stroke colour matches the land fill. */
.zone-label{
  font-family:"Fraunces",serif;font-weight:500;font-size:14px;
  fill:var(--ink);pointer-events:none;
  paint-order:stroke fill;
  stroke:var(--bg);stroke-width:3;stroke-linejoin:round;
}
.zone-count{
  font-size:10.5px;letter-spacing:.12em;text-transform:uppercase;
  fill:var(--muted);font-weight:700;pointer-events:none;
  paint-order:stroke fill;
  stroke:var(--bg);stroke-width:3;stroke-linejoin:round;
}
.zone-future{fill:color-mix(in srgb, var(--ink) 7%, var(--bg));stroke:var(--border);stroke-width:1;stroke-dasharray:3 4;opacity:.85}
.future-label{
  font-family:"Fraunces",serif;font-weight:500;font-size:16px;
  fill:var(--ink-soft);letter-spacing:.02em;
  paint-order:stroke fill;stroke:var(--bg);stroke-width:3;stroke-linejoin:round;
}
.future-sub{
  font-size:10px;letter-spacing:.18em;text-transform:uppercase;
  fill:var(--muted);font-weight:700;
  paint-order:stroke fill;stroke:var(--bg);stroke-width:3;stroke-linejoin:round;
}
@media(max-width:620px){
  .wales-map-wrap{padding:.6rem .6rem 1rem}
  .zone-label{font-size:12px}
  .zone-count{font-size:9px}
}

/* ─── Region tiles ───────────────────────────────────────── */
.regions{display:grid;grid-template-columns:repeat(4,1fr);gap:.9rem}
@media(max-width:920px){.regions{grid-template-columns:repeat(2,1fr)}}
@media(max-width:480px){.regions{grid-template-columns:1fr}}
.region-tile{
  position:relative;border-radius:12px;overflow:hidden;min-height:220px;
  color:var(--cream);padding:1.2rem;
  display:flex;flex-direction:column;justify-content:space-between;
  cursor:pointer;border:0;text-align:left;
  background:linear-gradient(135deg,var(--accent,#3d5a3b),#1b1f1a);
  transition:transform .3s cubic-bezier(.2,.7,.2,1),box-shadow .3s;
  isolation:isolate;
}
.region-tile::before{
  content:"";position:absolute;inset:0;z-index:-2;
  background-image:var(--img);background-size:cover;background-position:center;
  opacity:.62;transition:transform .6s;
}
.region-tile::after{
  content:"";position:absolute;inset:0;z-index:-1;
  background:linear-gradient(200deg,transparent 15%,rgba(27,31,26,.85) 92%);
}
.region-tile:hover{transform:translateY(-3px);box-shadow:var(--shadow-md)}
.region-tile:hover::before{transform:scale(1.06);opacity:.75}
.region-tile h4{font-family:"Fraunces",serif;font-weight:500;font-size:1.2rem;line-height:1.15}
.region-tile .r-tagline{font-size:.82rem;color:rgba(250,246,238,.85);margin-top:.5rem;line-height:1.4;max-width:22ch}
.region-tile .r-count{
  align-self:flex-end;font-family:"Fraunces",serif;font-size:1.7rem;font-weight:500;
  color:var(--amber);line-height:1;
}

/* ─── Finder ─────────────────────────────────────────────── */
.finder{background:var(--paper);padding:4rem 0 5rem;border-top:1px solid var(--border)}

.toolbar{
  background:var(--card);border:1px solid var(--border);border-radius:14px;
  padding:.7rem .85rem;margin-bottom:1.2rem;box-shadow:var(--shadow-sm);
  display:flex;gap:.6rem;align-items:center;flex-wrap:wrap;
}
.search-input{
  flex:1;min-width:220px;padding:.7rem .95rem .7rem 2.5rem;
  border:1px solid var(--border);border-radius:10px;background:var(--bg);
  font:inherit;font-size:.95rem;color:var(--ink);transition:border-color .15s,box-shadow .15s;
  background-image:url("data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%2376745f' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'><circle cx='11' cy='11' r='7'/><path d='m21 21-4.35-4.35'/></svg>");
  background-repeat:no-repeat;background-position:.9rem center;background-size:1rem;
}
.search-input:focus{outline:0;border-color:var(--moss-2);box-shadow:0 0 0 3px rgba(74,107,62,.15)}
.btn-ghost{background:transparent;color:var(--ink-soft);border:1px solid var(--border);
  padding:.6rem 1.05rem;border-radius:10px;font-weight:500;font-size:.85rem;cursor:pointer;transition:all .15s}
.btn-ghost:hover{border-color:var(--bracken);color:var(--bracken)}

details.filters{background:var(--card);border:1px solid var(--border);border-radius:14px;
  margin-bottom:1.5rem;box-shadow:var(--shadow-sm)}
details.filters>summary{list-style:none;cursor:pointer;padding:1rem 1.2rem;display:flex;
  justify-content:space-between;align-items:center;font-weight:600;font-size:.92rem;color:var(--ink)}
details.filters>summary::-webkit-details-marker{display:none}
details.filters>summary::after{content:"▾";color:var(--bracken);transition:transform .25s;font-size:.9rem}
details[open].filters>summary::after{transform:rotate(180deg)}
.f-body{padding:.3rem 1.2rem 1.3rem;display:grid;gap:1.1rem}
.f-group{display:flex;flex-direction:column;gap:.5rem}
.f-group .lbl{font-size:.7rem;letter-spacing:.16em;text-transform:uppercase;color:var(--muted);font-weight:700}
.chips{display:flex;flex-wrap:wrap;gap:.35rem}
.chip{padding:.35rem .8rem;border-radius:999px;background:var(--bg);color:var(--ink-soft);
  border:1px solid var(--border);cursor:pointer;font-size:.78rem;font-weight:500;
  user-select:none;transition:all .15s}
.chip:hover{border-color:var(--fern);color:var(--ink)}
.chip.on{background:var(--moss);color:var(--cream);border-color:var(--moss)}
.chip[data-chip-kind="difficulty"].on{background:var(--bracken);border-color:var(--bracken)}
.range-wrap{display:flex;align-items:center;gap:.7rem}
.range-wrap input[type=range]{flex:1;accent-color:var(--moss);height:3px}
.range-wrap .val{min-width:66px;text-align:right;font-variant-numeric:tabular-nums;font-size:.82rem;color:var(--muted);font-weight:600}
.two-col{display:grid;grid-template-columns:1fr 1fr;gap:1rem}
@media(max-width:620px){.two-col{grid-template-columns:1fr}}
.ck-list{display:grid;grid-template-columns:repeat(auto-fill,minmax(160px,1fr));gap:.4rem .9rem}
.ck-list label{display:flex;gap:.5rem;align-items:center;font-size:.85rem;cursor:pointer;color:var(--ink-soft)}
.ck-list input{accent-color:var(--moss);width:15px;height:15px;flex-shrink:0}
#sort{padding:.55rem .6rem;border:1px solid var(--border);border-radius:8px;
  background:var(--card);font:inherit;font-size:.86rem;color:var(--ink);cursor:pointer}

.result-bar{display:flex;justify-content:space-between;align-items:center;margin-bottom:1.1rem;
  font-size:.85rem;color:var(--muted);flex-wrap:wrap;gap:.4rem}
.result-bar #count{font-family:"Fraunces",serif;font-size:1.15rem;color:var(--ink);font-weight:500}

/* ─── Walk cards ─────────────────────────────────────────── */
#results{display:grid;gap:1rem;grid-template-columns:1fr}
@media(min-width:720px){#results{grid-template-columns:1fr 1fr}}
@media(min-width:1020px){#results{grid-template-columns:1fr 1fr 1fr}}
.card{
  background:var(--card);border:1px solid var(--border);border-radius:14px;
  padding:0;box-shadow:var(--shadow-sm);
  transition:transform .2s cubic-bezier(.2,.7,.2,1),box-shadow .2s,border-color .2s;
  display:flex;flex-direction:column;overflow:hidden;
}
.card:hover{transform:translateY(-3px);box-shadow:var(--shadow-md);border-color:var(--border-soft)}
.card-accent{height:5px;background:var(--accent,var(--moss))}
.card-body{padding:1.15rem 1.25rem 1.25rem;display:flex;flex-direction:column;gap:.65rem;flex:1}
.card-top{display:flex;gap:.5rem;align-items:flex-start;justify-content:space-between;flex-wrap:wrap}
.card h3{margin:0;font-family:"Fraunces",serif;font-weight:500;font-size:1.15rem;line-height:1.2;max-width:19ch;font-variation-settings:"opsz" 72}
.diff-pill{
  display:inline-flex;align-items:center;padding:.24rem .7rem;
  border-radius:999px;font-size:.68rem;font-weight:700;letter-spacing:.1em;
  text-transform:uppercase;color:var(--cream);white-space:nowrap;
}
.diff-Easy{background:#5a7a4e}
.diff-Easy\/Moderate{background:#7a8a45}
.diff-Moderate{background:#c68b2f}
.diff-Moderate\/Hard{background:#b65a2a}
.diff-Hard{background:#9a3c16}
.diff-Very\.Hard{background:#6b1a0a}

.region-row{display:flex;align-items:center;gap:.45rem;font-size:.7rem;color:var(--muted);
  letter-spacing:.14em;text-transform:uppercase;font-weight:700}
.region-row .dot{width:8px;height:8px;border-radius:50%;background:var(--accent,var(--moss));flex-shrink:0}

.stats-inline{display:grid;grid-template-columns:repeat(4,1fr);gap:.5rem;padding:.2rem 0}
.si{display:flex;flex-direction:column;gap:.05rem}
.si-val{font-family:"Fraunces",serif;font-weight:500;font-size:1.05rem;color:var(--ink);line-height:1.1;font-variation-settings:"opsz" 24}
.si-lbl{font-size:.63rem;letter-spacing:.1em;text-transform:uppercase;color:var(--muted);font-weight:700}

.feat-line{font-size:.88rem;color:var(--ink-soft);line-height:1.5;display:-webkit-box;
  -webkit-line-clamp:3;-webkit-box-orient:vertical;overflow:hidden;margin:.1rem 0}

.tag-row{display:flex;flex-wrap:wrap;gap:.3rem}
.tag{padding:.16rem .55rem;border-radius:4px;background:var(--bg);color:var(--ink-soft);
  font-size:.7rem;font-weight:500;border:1px solid var(--border)}

.card-btns{display:flex;gap:.5rem;margin-top:auto;padding-top:.3rem;flex-wrap:wrap}
.btn{flex:1;text-align:center;padding:.62rem .7rem;border-radius:10px;font-size:.86rem;
  font-weight:600;text-decoration:none;cursor:pointer;border:1px solid transparent;
  display:inline-flex;align-items:center;justify-content:center;gap:.35rem;line-height:1.2;
  font-family:"Inter",sans-serif;transition:all .15s;
}
.btn-primary{background:var(--slate);color:var(--cream);border-color:var(--slate)}
.btn-primary:hover{background:var(--bracken);border-color:var(--bracken)}
.btn-secondary{background:transparent;color:var(--ink);border-color:var(--border)}
.btn-secondary:hover{border-color:var(--bracken);color:var(--bracken)}

details.more{font-size:.85rem;margin-top:.2rem;flex-basis:100%}
details.more summary{display:none}
details.more dl{display:grid;grid-template-columns:auto 1fr;gap:.4rem .9rem;margin:.9rem 0 0;
  padding-top:.9rem;border-top:1px dashed var(--border)}
details.more dt{font-weight:700;color:var(--muted);font-size:.68rem;letter-spacing:.1em;text-transform:uppercase}
details.more dd{margin:0;color:var(--ink-soft);font-size:.84rem;line-height:1.5}

.empty{padding:3rem 1rem;text-align:center;background:var(--card);
  border-radius:14px;border:1px dashed var(--border);color:var(--muted);
  grid-column:1/-1}
.empty h4{font-family:"Fraunces",serif;color:var(--ink);font-size:1.2rem;margin-bottom:.5rem;font-weight:500}

/* ─── Welcome (pre-filter) state ─────────────────────────── */
.welcome{padding:3.5rem 1.5rem;text-align:center;background:var(--card);
  border-radius:14px;border:1px solid var(--border);color:var(--muted);
  grid-column:1/-1;box-shadow:var(--shadow-sm);position:relative;overflow:hidden}
.welcome::before{
  content:"";position:absolute;inset:0;z-index:0;opacity:.09;
  background-image:url("https://images.unsplash.com/photo-1441974231531-c6227db76b6e?auto=format&fit=crop&w=1600&q=60");
  background-size:cover;background-position:center;
}
.welcome>*{position:relative;z-index:1}
.welcome h4{font-family:"Fraunces",serif;color:var(--ink);font-size:1.55rem;margin-bottom:.6rem;font-weight:500;font-variation-settings:"opsz" 72}
.welcome p{max-width:44ch;margin:0 auto;line-height:1.6;font-size:.95rem}
.welcome-chips{display:flex;flex-wrap:wrap;gap:.5rem;justify-content:center;margin-top:1.5rem}
.welcome-chips button{background:var(--card);color:var(--ink-soft);border:1px solid var(--border);
  padding:.5rem 1.05rem;border-radius:999px;font-size:.82rem;font-weight:500;cursor:pointer;
  transition:all .15s;font-family:"Inter",sans-serif}
.welcome-chips button:hover{border-color:var(--bracken);color:var(--bracken);background:var(--paper)}

/* ─── Landscape band (between regions and finder) ─────────── */
.landscape-band{
  position:relative;min-height:240px;display:flex;align-items:center;justify-content:center;
  background-color:var(--moss);background-size:cover;background-position:center 55%;
  border-top:1px solid var(--moss);border-bottom:1px solid var(--moss);
  overflow:hidden;isolation:isolate;
}
.landscape-band::before{
  content:"";position:absolute;inset:0;z-index:-1;
  background:linear-gradient(180deg,rgba(27,31,26,.25) 0%,rgba(27,31,26,.55) 100%);
}
.band-quote{
  max-width:38ch;padding:3.5rem 1.25rem;text-align:center;
  font-family:"Fraunces",serif;font-weight:400;font-style:italic;
  font-size:clamp(1.1rem,2.3vw,1.65rem);color:var(--cream);line-height:1.35;
  font-variation-settings:"opsz" 96;
  text-shadow:0 1px 18px rgba(0,0,0,.35);
}
.band-quote small{display:block;font-style:normal;font-family:"Inter",sans-serif;
  font-size:.7rem;letter-spacing:.22em;text-transform:uppercase;
  color:var(--amber);margin-top:1rem;font-weight:700}
@media(max-width:720px){
  .landscape-band{min-height:200px}
  .band-quote{padding:2.6rem 1.25rem}
}

/* ─── Walk detail gallery and map ────────────────────────── */
.walk-gallery{
  display:grid;grid-template-columns:1fr 1fr 1fr;gap:.4rem;
  margin:.9rem 0 .2rem;
}
.walk-gallery img{
  width:100%;aspect-ratio:4/3;object-fit:cover;
  border-radius:8px;background:var(--stone);
  transition:transform .35s cubic-bezier(.2,.7,.2,1);
  box-shadow:var(--shadow-sm);
}
.walk-gallery img:hover{transform:scale(1.03)}
.walk-map{
  margin:.8rem 0 .25rem;border-radius:10px;overflow:hidden;
  border:1px solid var(--border);aspect-ratio:16/9;background:var(--stone);
  box-shadow:var(--shadow-sm);
}
.walk-map iframe{width:100%;height:100%;border:0;display:block}
.walk-map-links{display:flex;gap:1rem;margin:.2rem 0 .5rem;flex-wrap:wrap}
.walk-map-links a{font-size:.72rem;letter-spacing:.1em;text-transform:uppercase;
  color:var(--bracken);font-weight:700;padding:.2rem 0}
.walk-map-links a:hover{color:var(--bracken-dark);text-decoration:underline}
.walk-credits{font-size:.7rem;color:var(--muted);margin:.3rem 0 .6rem;line-height:1.4}
.walk-credits a{color:var(--muted);text-decoration:underline;text-decoration-color:var(--stone)}
.walk-credits a:hover{color:var(--bracken)}
.walk-credits strong{color:var(--ink-soft);font-weight:600}

/* ─── Ratings widget ─────────────────────────────────────── */
.walk-ratings{margin:1rem 0 .3rem;padding:.9rem 1rem;border:1px solid var(--stone);border-radius:10px;background:rgba(255,255,255,.55)}
.r-summary{display:flex;align-items:center;gap:.6rem;margin-bottom:.55rem;font-size:.88rem}
.r-summary strong{font-weight:600;color:var(--ink-soft)}
.r-stars{font-size:1.05rem;letter-spacing:.08em;color:var(--bracken);line-height:1}
.r-stars-empty{color:var(--stone)}
.r-count{font-size:.8rem;color:var(--muted);margin-left:.2rem}
.r-notice{font-size:.78rem;color:var(--muted);font-style:italic}
.r-hint{font-size:.82rem;color:var(--muted);margin:.1rem 0 .5rem}
.r-form .r-input, .r-form .r-comment, .r-form .r-actions{margin:.35rem 0}
.r-input{display:flex;align-items:center;gap:.35rem;flex-wrap:wrap}
.r-input-label{display:inline-block;min-width:9rem;font-size:.82rem;color:var(--ink-soft);font-weight:500}
.r-star{background:none;border:0;cursor:pointer;font-size:1.15rem;line-height:1;padding:2px 3px;color:var(--stone);transition:color .15s}
.r-star.r-on,.r-star:hover{color:var(--bracken)}
.r-star:focus-visible{outline:2px solid var(--amber);outline-offset:2px;border-radius:3px}
.r-comment{display:flex;flex-direction:column;gap:.25rem;margin-top:.55rem}
.r-comment span{font-size:.78rem;color:var(--muted)}
.r-comment em{font-style:italic;color:var(--muted)}
.r-comment textarea{font:inherit;font-size:.85rem;padding:.5rem .6rem;border:1px solid var(--stone);border-radius:6px;resize:vertical;min-height:60px;background:#fff}
.r-actions{display:flex;align-items:center;gap:.6rem;margin-top:.5rem;flex-wrap:wrap}
.r-save{background:var(--moss);color:#fff;border:0;padding:.45rem .95rem;border-radius:6px;font-size:.82rem;font-weight:600;cursor:pointer;letter-spacing:.05em}
.r-save:hover{background:var(--moss-dark)}
.r-signout{background:none;border:0;color:var(--muted);font-size:.78rem;cursor:pointer;text-decoration:underline}
.r-signout:hover{color:var(--ink-soft)}
.r-status{font-size:.78rem;color:var(--moss-dark);font-style:italic}
.r-signin .r-email{display:flex;gap:.4rem;flex-wrap:wrap}
.r-signin input[type=email]{flex:1 1 14rem;font:inherit;font-size:.85rem;padding:.45rem .6rem;border:1px solid var(--stone);border-radius:6px;background:#fff}
.r-signin button{background:var(--bracken);color:#fff;border:0;padding:.45rem .95rem;border-radius:6px;font-size:.82rem;font-weight:600;cursor:pointer;white-space:nowrap}
.r-signin button:hover{background:var(--bracken-dark)}

/* ─── Footer ─────────────────────────────────────────────── */
footer.site{background:var(--slate);color:var(--cream);padding:3.5rem 0 2rem;margin-top:0}
.foot-grid{display:grid;grid-template-columns:1.7fr 1fr 1fr;gap:2.5rem}
@media(max-width:720px){.foot-grid{grid-template-columns:1fr;gap:2rem}}
.foot-brand{display:flex;align-items:center;gap:.8rem;margin-bottom:1rem;color:var(--amber)}
.foot-brand-text{font-family:"Fraunces",serif;font-size:1.15rem;font-weight:500;color:var(--cream)}
.foot-brand-sub{font-size:.66rem;letter-spacing:.22em;text-transform:uppercase;color:rgba(250,246,238,.55);font-weight:600;margin-top:2px}
.foot-tagline{font-size:.92rem;color:rgba(250,246,238,.78);max-width:40ch;line-height:1.6}
.foot-col h5{font-family:"Fraunces",serif;font-weight:500;color:var(--amber);
  font-size:.78rem;text-transform:uppercase;letter-spacing:.16em;margin:0 0 .9rem}
.foot-col a,.foot-col span{display:block;font-size:.88rem;color:rgba(250,246,238,.78);
  padding:.22rem 0;transition:color .15s}
.foot-col a:hover{color:var(--amber)}
.foot-legal{margin-top:2.8rem;padding-top:1.3rem;border-top:1px solid rgba(250,246,238,.12);
  font-size:.76rem;color:rgba(250,246,238,.5);display:flex;justify-content:space-between;flex-wrap:wrap;gap:.8rem;letter-spacing:.02em}
</style>
</head>
<body>

<nav class="site">
  <div class="container nav-row">
    <a class="brand" href="#top">
      __LOGO_SVG__
      <div>
        <div class="brand-name"><b>South Wales</b> Walks</div>
        <span class="brand-sub">Curated · Wild · Walkable</span>
      </div>
    </a>
    <div class="nav-links">
      <a href="#regions">Regions</a>
      <a href="#finder">Find a walk</a>
      <a href="#about">About</a>
      <a class="nav-cta" href="#finder">Explore →</a>
    </div>
  </div>
</nav>

<header class="hero" id="top">
  <div class="container hero-inner">
    <span class="hero-eyebrow"><span class="dot"></span>__WALK_COUNT__ curated walks · 8 regions</span>
    <h1>Find your next<br><em>wild walk</em> across South Wales.</h1>
    <p class="hero-lede">Hand-compiled routes through the Brecon Beacons, Gower cliffs, Pembrokeshire coves, Wye Valley woods, Mid-Welsh reservoirs and the Forest of Dean. Filter by distance, difficulty, dog policy, waterfalls, pubs — then head out.</p>
    <div class="hero-meta">
      <span><b>__WALK_COUNT__</b>Walks</span>
      <span><b>8</b>Regions</span>
      <span><b>__NEAR_COUNT__</b>Near Monmouth</span>
    </div>
    <a class="hero-cta" href="#finder">Browse all walks ↓</a>
  </div>
</header>

<section class="stats">
  <div class="container stats-row">
    <div class="stat"><span class="stat-n">__WALK_COUNT__</span><span class="stat-l">Routes</span></div>
    <div class="stat"><span class="stat-n">8</span><span class="stat-l">Regions</span></div>
    <div class="stat"><span class="stat-n">__NEAR_COUNT__</span><span class="stat-l">Within 1hr of Monmouth</span></div>
    <div class="stat"><span class="stat-n">15</span><span class="stat-l">Points of interest</span></div>
  </div>
</section>

<section class="section">
  <div class="container">
    <div class="section-head">
      <div>
        <div class="section-eyebrow">Editor's picks</div>
        <h2 class="section-title"><em>Four</em> walks to start with.</h2>
      </div>
      <p class="section-sub">The first ones we'd send a visitor on: a summit, a coastline, a waterfall and a ruin.</p>
    </div>
    <div class="featured">__FEATURED_HTML__</div>
  </div>
</section>

<section class="section alt" id="regions">
  <div class="container">
    <div class="section-head">
      <div>
        <div class="section-eyebrow">Explore by region</div>
        <h2 class="section-title">From <em>coast</em> to <em>cantref</em>.</h2>
      </div>
      <p class="section-sub">Click a region on the map to filter — pick more than one if you like. Every walk was field-verified by a Monmouth local (well, their notebook).</p>
    </div>
    <div class="wales-map-wrap">__WALES_MAP_SVG__</div>
    <div class="regions">__REGION_TILES__</div>
  </div>
</section>

<div class="landscape-band" style="background-image:url('https://images.unsplash.com/photo-1470071459604-3b5ec3a7fe05?auto=format&amp;fit=crop&amp;w=2400&amp;q=65')">
  <p class="band-quote">Where the Usk meets the Wye, and the Beacons rise over the bracken — pick a line through the hills.<small>A note from the trail</small></p>
</div>

<section class="finder" id="finder">
  <div class="container">
    <div class="section-head">
      <div>
        <div class="section-eyebrow">Find your walk</div>
        <h2 class="section-title">All <em>__WALK_COUNT__</em> walks · yours to filter.</h2>
      </div>
      <p class="section-sub">Combine any filters — distance, difficulty, dog policy, waterfall obsession. Changes apply instantly.</p>
    </div>

    <div class="toolbar">
      <input class="search-input" id="search" type="search" placeholder="Search walk, pub, feature, village…" autocomplete="off" inputmode="search">
      <button class="btn-ghost" id="clear-all" type="button">Clear filters</button>
    </div>

    <details class="filters" open>
      <summary>Filters</summary>
      <div class="f-body">
        <div class="two-col">
          <div class="f-group">
            <span class="lbl">Drive from Monmouth NP25 3NT</span>
            <div class="range-wrap">
              <input type="range" id="drive-max" min="0" max="__MAX_DRIVE__" step="5" value="__MAX_DRIVE__">
              <span class="val" id="drive-val">Any</span>
            </div>
          </div>
          <div class="f-group">
            <span class="lbl">Max distance</span>
            <div class="range-wrap">
              <input type="range" id="dist-max" min="0" max="__MAX_MILES__" step="0.5" value="__MAX_MILES__">
              <span class="val" id="dist-val">Any</span>
            </div>
          </div>
        </div>

        <div class="f-group">
          <span class="lbl">Max elevation gain</span>
          <div class="range-wrap">
            <input type="range" id="elev-max" min="0" max="1000" step="50" value="1000">
            <span class="val" id="elev-val">Any</span>
          </div>
        </div>

        <div class="f-group">
          <span class="lbl">Difficulty</span>
          <div class="chips" id="difficulty-chips"></div>
        </div>

        <div class="f-group">
          <span class="lbl">Region</span>
          <div class="chips" id="region-chips"></div>
        </div>

        <div class="f-group">
          <span class="lbl">Points of interest</span>
          <div class="ck-list" id="poi-list"></div>
        </div>

        <div class="two-col">
          <div class="f-group">
            <span class="lbl">Dogs &amp; accessibility</span>
            <div class="ck-list">
              <label><input type="checkbox" id="dogs-yes"> Dogs allowed</label>
              <label><input type="checkbox" id="offlead"> Off-lead possible</label>
              <label><input type="checkbox" id="pushchair"> Pushchair friendly</label>
              <label><input type="checkbox" id="family"> Family friendly (Easy)</label>
            </div>
          </div>
          <div class="f-group">
            <span class="lbl">Sort by</span>
            <select id="sort">
              <option value="drive">Drive from Monmouth (closest first)</option>
              <option value="name">Name (A–Z)</option>
              <option value="miles-asc">Distance (shortest first)</option>
              <option value="miles-desc">Distance (longest first)</option>
              <option value="elev-asc">Elevation (easiest first)</option>
              <option value="elev-desc">Elevation (hardest first)</option>
            </select>
          </div>
        </div>
      </div>
    </details>

    <div class="result-bar">
      <span id="count">—</span>
      <span id="hint"></span>
    </div>

    <section id="results"></section>
  </div>
</section>

<footer class="site" id="about">
  <div class="container">
    <div class="foot-grid">
      <div>
        <div class="foot-brand">
          __LOGO_SVG_FOOTER__
          <div>
            <div class="foot-brand-text">South Wales Walks</div>
            <span class="foot-brand-sub">Cerddwr · Walking guide</span>
          </div>
        </div>
        <p class="foot-tagline">A hand-compiled, free, ad-free guide to the best walks of South Wales, Mid-Wales, the Marches and the Forest of Dean. Built in Monmouth. Maintained in the open on GitHub.</p>
      </div>
      <div class="foot-col">
        <h5>Navigate</h5>
        <a href="#top">Home</a>
        <a href="#regions">Regions</a>
        <a href="#finder">All walks</a>
      </div>
      <div class="foot-col">
        <h5>The small print</h5>
        <span>Verify tide tables, MoD range days, parking fees and pub opening times before setting out.</span>
        <a href="South_Wales_Walks_Database.xlsx">Download spreadsheet ↓</a>
      </div>
    </div>
    <div class="foot-legal">
      <span>© __YEAR__ Jason · South Wales Walks</span>
      <span>Data compiled 2026 · <em style="font-style:italic">Cerddwch yn ofalus</em> — walk safely.</span>
    </div>
  </div>
</footer>

<script>
const WALKS = __DATA_JSON__;
const REGIONS = __REGIONS_JSON__;
const DIFFICULTIES = __DIFFS_JSON__;
const TAGS = __TAGS_JSON__;
const REGION_ACCENT = __REGION_ACCENT_JSON__;
const REGION_SHORT  = __REGION_SHORT_JSON__;

const $ = q => document.querySelector(q);
const $$ = q => document.querySelectorAll(q);

function toggleDetails(btn){
  const d = btn.nextElementSibling;
  if (d && d.tagName === "DETAILS") d.open = !d.open;
}

function mapsUrl(w){
  const q = [w.postcode, w.name, "UK"].filter(Boolean).join(", ");
  return "https://www.google.com/maps?q=" + encodeURIComponent(q);
}

function esc(s){
  return String(s == null ? "" : s).replace(/[&<>"']/g, c =>
    ({"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;","'":"&#39;"}[c])
  );
}

function accentFor(region){ return REGION_ACCENT[region] || "#4a6b3e"; }
function shortFor(region){ return REGION_SHORT[region] || region; }

// Build the region-chip list using short names, with a lookup back to full names
const SHORTS = REGIONS.map(r => shortFor(r));
const SHORT_TO_FULL = Object.fromEntries(REGIONS.map(r => [shortFor(r), r]));

function makeChips(host, items, kind){
  host.innerHTML = items.map(v =>
    `<span class="chip" data-chip-kind="${kind}" data-v="${esc(v)}">${esc(v)}</span>`
  ).join("");
  host.querySelectorAll(".chip").forEach(el =>
    el.addEventListener("click", () => { el.classList.toggle("on"); apply(); })
  );
}
makeChips($("#difficulty-chips"), DIFFICULTIES, "difficulty");
makeChips($("#region-chips"), SHORTS, "region-short");

$("#poi-list").innerHTML = TAGS.map(t =>
  `<label><input type="checkbox" data-tag="${esc(t)}"> ${esc(t)}</label>`
).join("");
$$('#poi-list input').forEach(el => el.addEventListener("change", apply));

function bindRange(id, valId, fmt){
  const el = $("#"+id), v = $("#"+valId);
  const update = () => {
    const n = Number(el.value);
    v.textContent = (n >= Number(el.max)) ? "Any" : fmt(n);
    apply();
  };
  el.addEventListener("input", update);
  update();
}
bindRange("drive-max", "drive-val", n => n + " min");
bindRange("dist-max", "dist-val", n => n + " mi");
bindRange("elev-max", "elev-val", n => n + " m");

["dogs-yes","offlead","pushchair","family"].forEach(id =>
  $("#"+id).addEventListener("change", apply)
);
$("#sort").addEventListener("change", apply);
$("#search").addEventListener("input", apply);
$("#clear-all").addEventListener("click", reset);

// Region tiles → filter + scroll to finder
$$('.region-tile').forEach(t => t.addEventListener('click', () => {
  const wanted = t.dataset.regionShort;
  $$('.chip[data-chip-kind="region-short"]').forEach(c => {
    c.classList.toggle("on", c.dataset.v === wanted);
  });
  syncMapFromChips();
  apply();
  document.getElementById("finder").scrollIntoView({behavior:"smooth"});
}));

// Wales map zones → multi-select toggle, mirror to chips, scroll to finder.
// Each region now has TWO groups in the SVG (one for the shape inside the
// clip path, one for the label outside it) — both share the same
// data-region-full and we toggle .on on BOTH so the highlight + label
// inversion stay in sync.
function zoneToShort(full){
  return (REGION_SHORT && REGION_SHORT[full]) || full;
}
function setZoneOn(full, on){
  $$('.wales-zone-group').forEach(g => {
    if (g.dataset.regionFull === full) g.classList.toggle("on", on);
  });
}
function syncMapFromChips(){
  const active = new Set(selected("region-short"));
  $$('.wales-zone-group').forEach(g => {
    const short = zoneToShort(g.dataset.regionFull);
    g.classList.toggle("on", active.has(short));
  });
}
// Click handlers attach only to the shape groups (those have the path).
$$('.wales-zone-shape').forEach(g => {
  const path = g.querySelector('.wales-zone');
  const toggle = () => {
    const full  = g.dataset.regionFull;
    const short = zoneToShort(full);
    const chip  = document.querySelector(
      `.chip[data-chip-kind="region-short"][data-v="${CSS.escape(short)}"]`
    );
    if (!chip) return;
    chip.classList.toggle("on");
    setZoneOn(full, chip.classList.contains("on"));
    apply();
    document.getElementById("finder").scrollIntoView({behavior:"smooth"});
  };
  if (path){
    path.addEventListener("click", toggle);
    path.addEventListener("keydown", e => {
      if (e.key === "Enter" || e.key === " "){ e.preventDefault(); toggle(); }
    });
  }
});

// Keep map zones in sync when chips are clicked directly
$$('.chip[data-chip-kind="region-short"]').forEach(c =>
  c.addEventListener("click", () => setTimeout(syncMapFromChips, 0))
);

// Featured cards → search by walk name, scroll to finder, auto-open details
$$('.feat').forEach(card => card.addEventListener('click', () => {
  const name = card.dataset.walkName;
  if (!name) return;
  reset();
  $("#search").value = name;
  apply();
  document.getElementById("finder").scrollIntoView({behavior:"smooth"});
  setTimeout(() => {
    const target = [...document.querySelectorAll('#results .card')]
      .find(c => c.querySelector('h3')?.textContent === name);
    if (target){
      const det = target.querySelector('details.more');
      if (det) det.open = true;
      target.scrollIntoView({behavior:"smooth", block:"center"});
    }
  }, 400);
}));

function reset(){
  $("#search").value = "";
  $$(".chip.on").forEach(c => c.classList.remove("on"));
  $$(".wales-zone-group.on").forEach(g => g.classList.remove("on"));
  $$('#poi-list input:checked').forEach(c => c.checked = false);
  ["dogs-yes","offlead","pushchair","family"].forEach(id => $("#"+id).checked = false);
  $("#drive-max").value = $("#drive-max").max;
  $("#dist-max").value = $("#dist-max").max;
  $("#elev-max").value = $("#elev-max").max;
  ["drive-val","dist-val","elev-val"].forEach(i => $("#"+i).textContent = "Any");
  $("#sort").value = "drive";
  apply();
}

function selected(kind){
  return [...$$(`.chip.on[data-chip-kind="${kind}"]`)].map(c => c.dataset.v);
}
function selectedTags(){
  return [...$$('#poi-list input:checked')].map(c => c.dataset.tag);
}

function apply(){
  const q = $("#search").value.trim().toLowerCase();
  const diffs = selected("difficulty");
  const regShort = selected("region-short");
  const regs = regShort.map(s => SHORT_TO_FULL[s] || s);
  const tags  = selectedTags();
  const maxDrive = Number($("#drive-max").value);
  const maxMiles = Number($("#dist-max").value);
  const maxElev  = Number($("#elev-max").value);
  const needDogs = $("#dogs-yes").checked;
  const needOff  = $("#offlead").checked;
  const needPush = $("#pushchair").checked;
  const needFam  = $("#family").checked;
  const sort     = $("#sort").value;

  let out = WALKS.filter(w => {
    if (q){
      const hay = [w.name,w.region,w.sub,w.town,w.terrain,w.features,w.poi,
                   w.views,w.water,w.picnic,w.food,w.parking,w.notes,
                   (w.tags||[]).join(" ")].join(" ").toLowerCase();
      if (!hay.includes(q)) return false;
    }
    if (diffs.length && !diffs.includes(w.difficulty)) return false;
    if (regs.length  && !regs.includes(w.region))     return false;
    if (tags.length){
      const tset = new Set(w.tags || []);
      if (!tags.every(t => tset.has(t))) return false;
    }
    if (maxDrive < Number($("#drive-max").max) && (w.drive ?? 999) > maxDrive) return false;
    if (maxMiles < Number($("#dist-max").max) && (w.miles  ?? 999) > maxMiles) return false;
    if (maxElev  < Number($("#elev-max").max) && (w.elev   ?? 999) > maxElev)  return false;
    if (needDogs && String(w.dogs).toLowerCase() !== "yes") return false;
    if (needOff){
      const p = String(w.leash).toLowerCase();
      if (!(p.includes("off-lead") || p.includes("off lead"))) return false;
    }
    if (needPush){
      const p = String(w.pushchair).toLowerCase();
      if (!(p === "yes" || p.startsWith("partial") || p.includes("yes"))) return false;
    }
    if (needFam && !(String(w.difficulty).toLowerCase().startsWith("easy"))) return false;
    return true;
  });

  const keys = {
    "drive":      (a,b) => (a.drive||999) - (b.drive||999),
    "name":       (a,b) => a.name.localeCompare(b.name),
    "miles-asc":  (a,b) => (a.miles||0) - (b.miles||0),
    "miles-desc": (a,b) => (b.miles||0) - (a.miles||0),
    "elev-asc":   (a,b) => (a.elev||0)  - (b.elev||0),
    "elev-desc":  (a,b) => (b.elev||0)  - (a.elev||0),
  };
  out.sort(keys[sort] || keys["drive"]);
  render(out);
}

function walkCard(w){
  const diffClass = "diff-" + (w.difficulty || "").replace(/\s/g,".");
  const tagsHtml = (w.tags || []).slice(0,5).map(t => `<span class="tag">${esc(t)}</span>`).join("");
  const accent = accentFor(w.region);
  const short  = shortFor(w.region);
  const imgs = (w.images || []).slice(0,3);
  const credits = (w.photo_credits || []).slice(0,3);
  const galleryHtml = imgs.length
    ? `<div class="walk-gallery">${imgs.map((src,i) =>
        `<img loading="lazy" src="${esc(src)}" alt="${esc(w.name)} — scenery ${i+1}" onerror="this.style.visibility='hidden'">`
      ).join("")}</div>`
    : "";
  const creditHtml = credits.length
    ? `<div class="walk-credits">Photos: ${credits.map(c =>
        `<a href="${esc(c.page_url)}" target="_blank" rel="noopener noreferrer">${esc(c.title || 'photo')}</a> &copy; <strong>${esc(c.photographer)}</strong>`
      ).join(" · ")} &middot; <a href="${esc(credits[0].license_url)}" target="_blank" rel="noopener noreferrer">${esc(credits[0].license)}</a>, via ${esc(credits[0].source)}.</div>`
    : "";
  const mapQ = encodeURIComponent([w.postcode, w.name, "UK"].filter(Boolean).join(", "));
  const mapEmbed = w.postcode
    ? `<div class="walk-map"><iframe loading="lazy" src="https://maps.google.com/maps?q=${mapQ}&t=&z=13&ie=UTF8&iwloc=&output=embed" referrerpolicy="no-referrer-when-downgrade" title="Map of ${esc(w.name)}"></iframe></div>
       <div class="walk-map-links">
         <a href="https://explore.osmaps.com/search?q=${encodeURIComponent(w.postcode)}" target="_blank" rel="noopener noreferrer">Open in OS Maps ↗</a>
         <a href="https://www.openstreetmap.org/search?query=${encodeURIComponent(w.postcode)}" target="_blank" rel="noopener noreferrer">Open in OpenStreetMap ↗</a>
       </div>`
    : "";
  return `
    <article class="card" style="--accent:${accent}">
      <div class="card-accent"></div>
      <div class="card-body">
        <div class="region-row"><span class="dot"></span>${esc(short)}${w.sub ? " · " + esc(w.sub) : ""}</div>
        <div class="card-top">
          <h3>${esc(w.name)}</h3>
          <span class="diff-pill ${diffClass}">${esc(w.difficulty)}</span>
        </div>
        <div class="stats-inline">
          <div class="si"><span class="si-val">${w.miles}</span><span class="si-lbl">miles</span></div>
          <div class="si"><span class="si-val">${w.elev}</span><span class="si-lbl">m ascent</span></div>
          <div class="si"><span class="si-val">${w.time}</span><span class="si-lbl">hours</span></div>
          <div class="si"><span class="si-val">${w.drive ?? '?'}</span><span class="si-lbl">min drive</span></div>
        </div>
        <p class="feat-line">${esc(w.features || "")}</p>
        <div class="tag-row">${tagsHtml}</div>
        <div class="card-btns">
          <a class="btn btn-primary" href="${mapsUrl(w)}" target="_blank" rel="noopener noreferrer">Directions ↗</a>
          <button class="btn btn-secondary" type="button" onclick="toggleDetails(this)">More detail</button>
          <details class="more">
            <summary>Details</summary>
            ${galleryHtml}
            ${creditHtml}
            ${mapEmbed}
            <dl>
              <dt>Start</dt><dd>${esc(w.parking || '—')} (${esc(w.postcode || '—')})</dd>
              <dt>Terrain</dt><dd>${esc(w.terrain || '—')}</dd>
              <dt>Route</dt><dd>${esc(w.route || '—')}</dd>
              <dt>Dogs</dt><dd>${esc(w.dogs)} · ${esc(w.leash || '')}</dd>
              <dt>Pushchair</dt><dd>${esc(w.pushchair || '—')}</dd>
              <dt>Waymarked</dt><dd>${esc(w.waymarked || '—')}</dd>
              <dt>Best season</dt><dd>${esc(w.season || '—')}</dd>
              <dt>Points of interest</dt><dd>${esc(w.poi || '—')}</dd>
              <dt>Viewpoints</dt><dd>${esc(w.views || '—')}</dd>
              <dt>Water features</dt><dd>${esc(w.water || '—')}</dd>
              <dt>Picnic spots</dt><dd>${esc(w.picnic || '—')}</dd>
              <dt>Food &amp; drink</dt><dd>${esc(w.food || '—')}</dd>
              <dt>Toilets</dt><dd>${esc(w.toilets || '—')}</dd>
              <dt>Public transport</dt><dd>${esc(w.transport || '—')}</dd>
              <dt>Hazards / notes</dt><dd>${esc(w.notes || '—')}</dd>
            </dl>
            <div class="walk-ratings" data-ratings-for="${w.id}"></div>
          </details>
        </div>
      </div>
    </article>
  `;
}

function hasActiveFilter(){
  const q = $("#search").value.trim();
  if (q) return true;
  if (document.querySelector(".chip.on")) return true;
  if (document.querySelector("#poi-list input:checked")) return true;
  if (["dogs-yes","offlead","pushchair","family"].some(id => $("#"+id).checked)) return true;
  const atMax = id => Number($("#"+id).value) >= Number($("#"+id).max);
  if (!atMax("drive-max") || !atMax("dist-max") || !atMax("elev-max")) return true;
  return false;
}

function quickFilter(kind){
  if (kind === 'near'){
    const el = $("#drive-max");
    el.value = Math.min(60, Number(el.max));
    el.dispatchEvent(new Event('input'));
    return;
  }
  const chk = document.querySelector(`#poi-list input[data-tag="${kind}"]`);
  if (chk){ chk.checked = true; apply(); return; }
  const chip = [...document.querySelectorAll('.chip[data-chip-kind="difficulty"]')].find(c => c.dataset.v === kind);
  if (chip){ chip.classList.add("on"); apply(); return; }
}

function render(list){
  const host = $("#results");
  const active = hasActiveFilter();
  if (!active){
    host.innerHTML = `
      <div class="welcome">
        <h4>Search or filter to see walks.</h4>
        <p>All ${WALKS.length} walks are ready to explore. Search for a place, tap a region tile above, or try a quick start:</p>
        <div class="welcome-chips">
          <button type="button" onclick="quickFilter('near')">Within 1hr of Monmouth</button>
          <button type="button" onclick="quickFilter('Family Friendly')">Family friendly</button>
          <button type="button" onclick="quickFilter('Waterfall')">Waterfalls</button>
          <button type="button" onclick="quickFilter('Beach / Coastal')">Coastal</button>
          <button type="button" onclick="quickFilter('Mountain / Summit')">Mountain summits</button>
          <button type="button" onclick="quickFilter('Castle')">Castles</button>
          <button type="button" onclick="quickFilter('Pub / Cafe on Route')">With a pub on route</button>
        </div>
      </div>`;
    $("#count").textContent = `${WALKS.length} walks indexed`;
    $("#hint").textContent = "filter to reveal";
    return;
  }
  if (!list.length){
    host.innerHTML = `<div class="empty"><h4>No walks match these filters.</h4>Try widening a range or unticking a point of interest.</div>`;
  } else {
    host.innerHTML = list.map(walkCard).join("");
  }
  $("#count").textContent = `${list.length} walk${list.length === 1 ? '' : 's'}`;
  $("#hint").textContent = list.length === WALKS.length
    ? "showing all"
    : `of ${WALKS.length}`;
  // Re-paint ratings widgets after every filter/search — apply() replaces
  // #results.innerHTML, so every card that appears gets a fresh empty
  // .walk-ratings container. Without this call, only the cards rendered
  // at init-time ever show the ratings summary and sign-in form.
  if (window.ratings && window.ratings.refreshAll) window.ratings.refreshAll();
}

apply();

// Ratings widget — loads after the walk list so every details block has
// a <div data-ratings-for="{id}"> slot ready.
// Placeholders __INJECT_SUPABASE_URL__ / __INJECT_SUPABASE_ANON_KEY__
// are substituted at build time by build_gui.py. The VARIABLE NAMES
// (window.__SUPABASE_URL__ etc.) must NOT match the placeholder tokens,
// otherwise Python's str.replace() rewrites them too and we get
// `window. = "…"` — which is a syntax error that kills the whole script.
window.__SUPABASE_URL__      = "__INJECT_SUPABASE_URL__";
window.__SUPABASE_ANON_KEY__ = "__INJECT_SUPABASE_ANON_KEY__";
</script>
<script src="ratings.js"></script>
<script>
if (window.ratings) window.ratings.init();
</script>
</body>
</html>
"""

HTML = (HTML
    .replace("__LOGO_SVG__", LOGO_SVG)
    .replace("__LOGO_SVG_FOOTER__", LOGO_SVG_FOOTER)
    .replace("__WALK_COUNT__", str(len(data)))
    .replace("__NEAR_COUNT__", str(near_count))
    .replace("__MAX_DRIVE__", str(int(max_drive)))
    .replace("__MAX_MILES__", str(float(max_miles)))
    .replace("__YEAR__", str(datetime.date.today().year))
    .replace("__FEATURED_HTML__", featured_html)
    .replace("__REGION_TILES__", region_tiles_html)
    .replace("__WALES_MAP_SVG__", wales_map_html)
    .replace("__DATA_JSON__", json.dumps(data, ensure_ascii=False))
    .replace("__REGIONS_JSON__", json.dumps(regions, ensure_ascii=False))
    .replace("__DIFFS_JSON__", json.dumps(difficulties, ensure_ascii=False))
    .replace("__TAGS_JSON__", json.dumps(all_tags, ensure_ascii=False))
    .replace("__REGION_ACCENT_JSON__", json.dumps(region_accent_js, ensure_ascii=False))
    .replace("__REGION_SHORT_JSON__", json.dumps(region_short_js, ensure_ascii=False))
    .replace("__INJECT_SUPABASE_URL__", SUPABASE_URL)
    .replace("__INJECT_SUPABASE_ANON_KEY__", SUPABASE_ANON_KEY)
)

with open(OUT, "w", encoding="utf-8") as f:
    f.write(HTML)

print(f"Wrote {OUT}")
print(f"  size: {len(HTML):,} chars")
print(f"  regions: {len(regions)}")
print(f"  tags: {len(all_tags)} ({', '.join(all_tags)})")
print(f"  featured walks: {len(featured_walks)}")
print(f"  region tiles: {len(region_tiles)}")
print(f"  near Monmouth (<=60min): {near_count}")
