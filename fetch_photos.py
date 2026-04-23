"""Pre-fetch Wikimedia Commons photos for every walk and cache them.

Wikimedia Commons has excellent UK landscape coverage (much of Geograph has
been bulk-imported there) with a reliable, free, well-documented API. For
each walk we:
  1. Convert the start postcode -> lat/lon via postcodes.io (free, no auth).
  2. Use Commons' `list=geosearch` to find image files near that point.
  3. Fetch `imageinfo` with extmetadata to get URLs, author, licence.
  4. Keep only real photos (image/jpeg), filter out maps/logos/diagrams,
     cache the top N in `photos_cache.json`.

Why Commons (not Geograph):
  The Geograph syndicator's public endpoint silently returns 0 results for
  `ll=<lat>,<lon>` queries even in photo-dense areas like the Brecon Beacons.
  Commons' `geosearch` is the same idea but actually works, and it exposes
  most Geograph photos anyway via the long-running Commons:Geograph import.

Attribution: required by CC-BY-SA / CC-BY. We always render the author with
a link back to the Commons file page.

Usage:
  python fetch_photos.py                # fill missing / empty entries
  python fetch_photos.py --refresh      # re-fetch everything
  python fetch_photos.py --walk "Pen y Fan Circular (Motorway Route)"
  python fetch_photos.py --verbose      # log every step

If Commons or postcodes.io is unreachable, the script logs a warning and
keeps any previous cached entries. build_gui.py renders an empty gallery
for walks with no cached photos, so the failure mode is "no photos" — never
"wrong photos".
"""
from __future__ import annotations
import argparse
import json
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).parent
XLSX = HERE / "South_Wales_Walks_Database.xlsx"
CACHE = HERE / "photos_cache.json"

# Wikimedia requires a descriptive User-Agent on anon API requests.
USER_AGENT = (
    "south-wales-walks/3.0 "
    "(+https://github.com/Edition74/south-wales-walks; edition74@outlook.com)"
)
COMMONS_API = "https://commons.wikimedia.org/w/api.php"
POSTCODES_IO = "https://api.postcodes.io/postcodes/"
# Short delay between postcodes.io + Commons calls. Both have no strict
# public limit but we stay polite.
DELAY_SECS = 0.3
REQ_TIMEOUT = 15
# Search radius in metres around each walk's start. 10 km covers the walk
# itself plus a generous ring of surrounding viewpoints.
RADIUS_M = 10_000
# How many candidate files to pull from Commons before filtering. We throw
# away maps/logos/diagrams, so ask for plenty of headroom before keeping 3.
CANDIDATES = 30
# Thumbnail width in pixels — Commons will scale server-side.
THUMB_WIDTH = 1024

# File-title substrings that almost certainly aren't scenic photos of the
# area. Checked case-insensitively on the raw "File:…" title.
EXCLUDE_TITLE_PARTS = (
    "coat of arms",
    "coat_of_arms",
    "flag of",
    "flag_of",
    "logo",
    "map of",
    "map_of",
    "location map",
    "location_map",
    "diagram",
    "schematic",
    "plan of",
    "plan_of",
    "road sign",
    "wikidata",
)


def _http_get(url: str, verbose: bool = False) -> bytes | None:
    req = urllib.request.Request(
        url, headers={"User-Agent": USER_AGENT, "Accept": "application/json"}
    )
    try:
        with urllib.request.urlopen(req, timeout=REQ_TIMEOUT) as r:
            return r.read()
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as e:
        if verbose:
            print(f"    ! request failed: {e}", file=sys.stderr)
        return None


def postcode_to_lat_lon(postcode: str, verbose: bool = False) -> tuple[float, float] | None:
    """Return (lat, lon) for a UK postcode, or None if not found."""
    if not postcode:
        return None
    encoded = urllib.parse.quote(postcode.strip())
    body = _http_get(f"{POSTCODES_IO}{encoded}", verbose=verbose)
    if not body:
        return None
    try:
        obj = json.loads(body)
    except json.JSONDecodeError:
        return None
    if obj.get("status") != 200:
        return None
    res = obj.get("result") or {}
    lat, lon = res.get("latitude"), res.get("longitude")
    if lat is None or lon is None:
        return None
    return (float(lat), float(lon))


def commons_geosearch(lat: float, lon: float, verbose: bool = False) -> list[str]:
    """Return a list of 'File:Foo.jpg' titles near (lat, lon). Empty on failure."""
    params = {
        "action":        "query",
        "list":          "geosearch",
        "gscoord":       f"{lat}|{lon}",
        "gsradius":      RADIUS_M,
        "gslimit":       CANDIDATES,
        "gsnamespace":   6,       # File: namespace
        "format":        "json",
        "formatversion": 2,
    }
    url = f"{COMMONS_API}?{urllib.parse.urlencode(params)}"
    if verbose:
        print(f"    commons geosearch: {url}")
    body = _http_get(url, verbose=verbose)
    if not body:
        return []
    try:
        obj = json.loads(body)
    except json.JSONDecodeError:
        return []
    results = obj.get("query", {}).get("geosearch", []) or []
    titles: list[str] = []
    for r in results:
        title = r.get("title") or ""
        if not title.startswith("File:"):
            continue
        low = title.lower()
        if any(part in low for part in EXCLUDE_TITLE_PARTS):
            continue
        titles.append(title)
    return titles


_TAG_RX = re.compile(r"<[^>]+>")


def _clean(s: str) -> str:
    return _TAG_RX.sub("", s or "").strip()


def commons_imageinfo(titles: list[str], verbose: bool = False) -> list[dict]:
    """Fetch image URLs + attribution for a batch of Commons file titles."""
    if not titles:
        return []
    params = {
        "action":              "query",
        "prop":                "imageinfo",
        "iiprop":              "url|extmetadata|mime|size",
        "iiurlwidth":          THUMB_WIDTH,
        "iiextmetadatafilter": "ImageDescription|Artist|LicenseShortName|LicenseUrl|Credit",
        "titles":              "|".join(titles),
        "format":              "json",
        "formatversion":       2,
    }
    url = f"{COMMONS_API}?{urllib.parse.urlencode(params)}"
    if verbose:
        print(f"    commons imageinfo: {len(titles)} title(s)")
    body = _http_get(url, verbose=verbose)
    if not body:
        return []
    try:
        obj = json.loads(body)
    except json.JSONDecodeError:
        return []

    # Preserve the geosearch ordering (distance-ranked) rather than Commons'
    # alphabetical page order.
    pages_by_title = {p.get("title"): p for p in obj.get("query", {}).get("pages", []) or []}
    out: list[dict] = []
    for t in titles:
        page = pages_by_title.get(t)
        if not page:
            continue
        ii = (page.get("imageinfo") or [None])[0]
        if not ii:
            continue
        mime = ii.get("mime", "")
        # JPEGs are overwhelmingly landscape photos on Commons. PNGs are
        # often screenshots/logos/maps; SVGs are diagrams; TIFFs are
        # digitised documents. Stick to JPEG for predictable results.
        if mime != "image/jpeg":
            continue

        meta = ii.get("extmetadata") or {}

        def mv(key: str) -> str:
            return (meta.get(key) or {}).get("value", "") or ""

        author = _clean(mv("Artist")) or "Wikimedia Commons contributor"
        credit = _clean(mv("Credit"))
        lic = _clean(mv("LicenseShortName")) or "See Commons"
        lic_url = _clean(mv("LicenseUrl"))
        desc = _clean(mv("ImageDescription"))
        # Fall back to the filename (without extension/namespace) when no
        # description exists.
        pretty_title = (
            desc[:140]
            if desc
            else t.removeprefix("File:").rsplit(".", 1)[0].replace("_", " ")
        )

        out.append({
            "title":        pretty_title,
            "page_url":     ii.get("descriptionurl", ""),
            "thumb":        ii.get("thumburl") or ii.get("url"),
            "url":          ii.get("url"),
            "photographer": author,
            "credit":       credit,
            "license":      lic,
            "license_url":  lic_url,
            "source":       "Wikimedia Commons",
        })
    return out


def read_walks() -> list[dict]:
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Walks"]
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    walks = []
    for r in range(2, ws.max_row + 1):
        rec = {h: ws.cell(r, c).value for h, c in headers.items()}
        if rec.get("ID") and rec.get("Start Postcode"):
            walks.append({
                "id":       rec["ID"],
                "name":     rec["Walk Name"],
                "postcode": rec["Start Postcode"],
            })
    return walks


def main() -> None:
    ap = argparse.ArgumentParser(description="Fetch Wikimedia Commons photos for every walk.")
    ap.add_argument("--refresh", action="store_true", help="Re-fetch all walks, even cached ones.")
    ap.add_argument("--walk", help="Only fetch a single walk by name.")
    ap.add_argument("--limit", type=int, default=3, help="Photos to keep per walk (default 3).")
    ap.add_argument("--verbose", "-v", action="store_true", help="Log each HTTP step.")
    ap.add_argument("--dry-run", action="store_true", help="Don't write the cache file.")
    args = ap.parse_args()

    if CACHE.exists():
        cache: dict = json.loads(CACHE.read_text(encoding="utf-8"))
    else:
        cache = {"version": 3, "walks": {}}
    # v1 used the broken Geograph `q=postcode` query; v2 used Geograph's
    # `ll=` which silently returns empty. v3 moves to Wikimedia Commons.
    # Either older version is force-cleared on first v3 run.
    if cache.get("version") != 3:
        print(f"cache version {cache.get('version')!r} < 3 — clearing (old source returned empties)")
        cache = {"version": 3, "walks": {}}
    cache.setdefault("walks", {})

    walks = read_walks()
    if args.walk:
        walks = [w for w in walks if w["name"] == args.walk]
        if not walks:
            print(f"No walk matches: {args.walk}", file=sys.stderr)
            sys.exit(2)

    fetched = skipped = failed = no_coords = 0
    for w in walks:
        wid = str(w["id"])
        entry = cache["walks"].get(wid, {})
        if not args.refresh and entry.get("photos"):
            skipped += 1
            continue
        print(f"- [{wid}] {w['name']}  ({w['postcode']})")

        coords = postcode_to_lat_lon(w["postcode"], verbose=args.verbose)
        if not coords:
            print(f"    ! postcodes.io couldn't resolve {w['postcode']}")
            no_coords += 1
            cache["walks"].setdefault(wid, {}).update({
                "last_attempt": int(time.time()),
                "last_status":  "no-coords",
            })
            time.sleep(DELAY_SECS)
            continue
        lat, lon = coords

        titles = commons_geosearch(lat, lon, verbose=args.verbose)
        if not titles:
            print(f"    ! commons returned 0 files within {RADIUS_M//1000}km of {lat:.4f},{lon:.4f}")
            failed += 1
            cache["walks"].setdefault(wid, {}).update({
                "last_attempt": int(time.time()),
                "last_status":  "empty-search",
            })
            time.sleep(DELAY_SECS)
            continue

        photos = commons_imageinfo(titles, verbose=args.verbose)
        if not photos:
            print(f"    ! all {len(titles)} candidates filtered out (non-photos)")
            failed += 1
            cache["walks"].setdefault(wid, {}).update({
                "last_attempt": int(time.time()),
                "last_status":  "empty-filter",
            })
            time.sleep(DELAY_SECS)
            continue

        keep = photos[: args.limit]
        cache["walks"][wid] = {
            "name":         w["name"],
            "postcode":     w["postcode"],
            "lat":          lat,
            "lon":          lon,
            "photos":       keep,
            "last_attempt": int(time.time()),
            "last_status":  "ok",
        }
        print(f"    ok — kept {len(keep)} of {len(photos)} photos, first: {keep[0]['title']!r}")
        fetched += 1
        time.sleep(DELAY_SECS)

    print(
        f"\nFetched: {fetched}   cached-already: {skipped}   "
        f"failed(empty): {failed}   failed(no-coords): {no_coords}"
    )
    if not args.dry_run:
        CACHE.write_text(json.dumps(cache, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"Wrote {CACHE}")


if __name__ == "__main__":
    main()
