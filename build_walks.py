"""Build South Wales Walks & Hikes Database (xlsx).

The xlsx is a DERIVED ARTIFACT. The source of truth is walks/*.json, loaded
here via walks_loader. Edit JSON files to change walk data; re-run this
script to regenerate the spreadsheet.

Columns are designed for filtering/sorting in Excel:
  numeric: distance, elevation, time
  categorical: region, difficulty, route type, terrain, dogs, etc.
  free text: highlights, points of interest, food & drink, etc.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from walks_loader import load_walks

# ---------------------------------------------------------------------------
# Column schema
# ---------------------------------------------------------------------------
COLUMNS = [
    ("ID", 6),
    ("Walk Name", 34),
    ("Region", 22),
    ("Sub-area", 22),
    ("Nearest Town", 18),
    ("Distance (mi)", 12),
    ("Distance (km)", 12),
    ("Elevation Gain (m)", 14),
    ("Est. Time (hrs)", 13),
    ("Difficulty", 12),
    ("Route Type", 14),
    ("Terrain", 28),
    ("Dogs Allowed", 14),
    ("Dog Lead Policy", 22),
    ("Pushchair Friendly", 16),
    ("Waymarked", 12),
    ("Best Season", 18),
    ("Key Features / Highlights", 50),
    ("Points of Interest", 40),
    ("Viewpoints & Beauty Spots", 40),
    ("Water Features", 30),
    ("Picnic Spots", 30),
    ("Parking & Start", 34),
    ("Food & Drink Nearby", 36),
    ("Toilets", 18),
    ("Public Transport", 26),
    ("Hazards / Notes", 34),
    ("Start Postcode", 14),
    ("Drive from Monmouth (mins)", 18),
    ("Map Link", 20),
]

# ---------------------------------------------------------------------------
# Walks data — 100 walks across South Wales
# Order of fields matches COLUMNS (except km, which is a formula col, so it is
# left blank here and populated in the write loop).
# ---------------------------------------------------------------------------
# Field order per entry:
# name, region, sub_area, town, miles, elev_m, time_h, difficulty, route,
# terrain, dogs, leash, pushchair, waymarked, season, features, poi,
# viewpoints, water, picnic, parking, food, toilets, transport, notes

BRECON = "Brecon Beacons / Bannau Brycheiniog"
GOWER = "Gower & Swansea Bay"
PEMBS = "Pembrokeshire (South)"
VALES = "Valleys & Vale of Glamorgan"
WYEMON = "Wye Valley & Monmouthshire"
MIDWAL = "Mid Wales (Powys & Ceredigion)"
CARMS = "Carmarthenshire & West Wales"
BORDERS = "English Borders (Forest of Dean & Herefordshire)"

# ---------------------------------------------------------------------------
# Walks data — loaded from walks/*.json (the source of truth). The old
# WALKS list of 25-tuples and POSTCODES dict are reconstructed here so the
# rendering code further down doesn't need to change.
# ---------------------------------------------------------------------------
def _load_walks_data():
    """Build (WALKS_tuples, POSTCODES_dict) from walks/*.json."""
    _walks_json = load_walks()
    _walks_tuples = []
    _postcodes = {}
    for w in _walks_json:
        _walks_tuples.append((
            w.get("name"),
            w.get("region"),
            w.get("sub_area"),
            w.get("nearest_town"),
            w.get("distance_mi"),
            w.get("elevation_gain_m"),
            w.get("est_time_hrs"),
            w.get("difficulty"),
            w.get("route_type"),
            w.get("terrain"),
            w.get("dogs_allowed"),
            w.get("dog_lead_policy"),
            w.get("pushchair_friendly"),
            w.get("waymarked"),
            w.get("best_season"),
            w.get("highlights"),
            w.get("points_of_interest"),
            w.get("viewpoints"),
            w.get("water_features"),
            w.get("picnic_spots"),
            w.get("parking_start"),
            w.get("food_drink_nearby"),
            w.get("toilets"),
            w.get("public_transport"),
            w.get("hazards_notes"),
        ))
        _postcodes[w.get("name")] = (
            w.get("start_postcode") or "",
            w.get("drive_from_monmouth_mins"),
        )
    return _walks_tuples, _postcodes


WALKS, POSTCODES = _load_walks_data()


def map_link(postcode: str) -> str:
    """Return a Google Maps URL for a given postcode.

    Uses the plain ?q= form, which works reliably on mobile browsers and
    opens inside the Google Maps app if installed.
    """
    from urllib.parse import quote
    return f"https://www.google.com/maps?q={quote(postcode + ' UK')}"


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
wb = Workbook()

# --- Walks sheet
ws = wb.active
ws.title = "Walks"

header_fill = PatternFill("solid", start_color="1F4E78")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
cell_font = Font(name="Arial", size=10)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Header row
for col_idx, (name, width) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Data rows — insert formula for km
REGION_COLOURS = {
    BRECON: "E8F1FA",
    GOWER: "E8F6E8",
    PEMBS: "FFF4E6",
    VALES: "F3E8F7",
    WYEMON: "FDECEC",
    MIDWAL: "E8F7F6",   # soft teal — mid-Wales rivers
    CARMS: "FFF9E0",    # soft gold — Tywi valley
    BORDERS: "EFEEE0",  # soft stone — Forest of Dean oak
}

for i, w in enumerate(WALKS, start=1):
    (name, region, subarea, town, miles, elev_m, time_h, diff, route, terrain,
     dogs, leash, pushchair, waymarked, season, features, poi, views, water,
     picnic, parking, food, toilets, transport, notes) = w
    row = i + 1
    row_fill = PatternFill("solid", start_color=REGION_COLOURS[region])

    meta = POSTCODES.get(name, ("", ""))
    postcode, drive_min = meta
    link = map_link(postcode) if postcode else ""

    values = [
        i, name, region, subarea, town, miles,
        None,  # km – formula below
        elev_m, time_h, diff, route, terrain,
        dogs, leash, pushchair, waymarked, season,
        features, poi, views, water, picnic,
        parking, food, toilets, transport, notes,
        postcode, drive_min, link,
    ]
    for col_idx, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=v)
        cell.font = cell_font
        cell.fill = row_fill
        if col_idx in (1, 6, 7, 8, 9, 10, 11, 13, 14, 15, 16, 28, 29):
            cell.alignment = center
        else:
            cell.alignment = left_wrap
    # Formula column G = miles * 1.609344 (km)
    ws.cell(row=row, column=7).value = f"=F{row}*1.609344"
    ws.cell(row=row, column=7).number_format = "0.0"
    ws.cell(row=row, column=6).number_format = "0.0"
    ws.cell(row=row, column=8).number_format = "0"
    ws.cell(row=row, column=9).number_format = "0.00"
    # Hyperlink the map link column and show as "Map" text
    if link:
        link_cell = ws.cell(row=row, column=30)
        link_cell.value = "Open in Google Maps"
        link_cell.hyperlink = link
        link_cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
    # Drive-time numeric format
    if isinstance(drive_min, int):
        ws.cell(row=row, column=29).number_format = "0"

# Freeze header, enable auto-filter
ws.freeze_panes = "C2"
last_col = get_column_letter(len(COLUMNS))
last_row = len(WALKS) + 1
ws.auto_filter.ref = f"A1:{last_col}{last_row}"

# Row heights to accommodate wrapped text
ws.row_dimensions[1].height = 32
for r in range(2, last_row + 1):
    ws.row_dimensions[r].height = 80

# ---------------------------------------------------------------------------
# Legend sheet
# ---------------------------------------------------------------------------
legend = wb.create_sheet("Legend & How to Use")
legend.column_dimensions['A'].width = 28
legend.column_dimensions['B'].width = 80

legend_rows = [
    ("South Wales Walks & Hikes — Database v1", ""),
    ("", ""),
    ("How to use", "Click the filter arrows in row 1 of the 'Walks' sheet to sort or filter by distance, difficulty, region, terrain, dog policy, etc. You can combine filters (e.g., Region = Brecon Beacons + Difficulty = Easy + Dogs Allowed = Yes)."),
    ("", ""),
    ("Regions", ""),
    (BRECON, "Central Beacons, Black Mountain (west), Black Mountains (east), Fforest Fawr, Waterfall Country, Llangorse Lake."),
    (GOWER, "Gower Peninsula plus Swansea Bay promenade, Mumbles and Clyne Valley."),
    (PEMBS, "Pembrokeshire coast from Amroth/Tenby round to Dale — focused on the southern coastline."),
    (VALES, "Cardiff, Vale of Glamorgan (Heritage Coast), Valleys (Rhondda, Cynon, Rhymney, Ebbw, Afan)."),
    (WYEMON, "Wye Valley from Chepstow to Monmouth, plus Monmouthshire castles and Mon+Brec canal."),
    (MIDWAL, "Mid Wales — Powys & Ceredigion: Elan Valley reservoirs, Plynlimon, Cambrian Mountains, Ceredigion coast and Lake Vyrnwy."),
    (CARMS, "Carmarthenshire & West Wales: Tywi Valley (Dinefwr, Botanic Garden, Dryslwyn), Carmarthen Bay (Pembrey, Kidwelly) and Teifi Valley."),
    (BORDERS, "English borders — Forest of Dean (Speech House, sculpture trail, Cannop) and Herefordshire (Black Hill, Golden Valley, Great Doward)."),
    ("", ""),
    ("Difficulty", ""),
    ("Easy", "Mostly flat or gentle gradients, typically under 5 miles, well-surfaced or firm paths. Suitable for most abilities."),
    ("Moderate", "Some ascent (150–400m), uneven paths, 4–8 miles. Requires a reasonable level of fitness and sensible footwear."),
    ("Hard", "Significant ascent (400m+), exposed terrain, 8+ miles, or demanding conditions. Navigation skills and proper kit recommended."),
    ("Very Hard", "Long mountain days, exposed ridges, remote terrain, 900m+ ascent. Experienced hillwalkers only."),
    ("", ""),
    ("Route type", ""),
    ("Loop", "Returns to start without retracing steps."),
    ("Out-and-back", "Walk out to a point and return the same way."),
    ("Linear", "Different start and finish — requires transport arrangement (bus/train/second car)."),
    ("", ""),
    ("Dogs Allowed / Dog Lead Policy", ""),
    ("Yes / No", "Whether dogs are permitted on the route at all."),
    ("On lead", "Lead required throughout — usually because of livestock, ground-nesting birds, cliffs, or seasonal restrictions."),
    ("Off-lead OK", "Dogs can be off lead on most of the route (still recall-trained)."),
    ("Seasonal beach bans", "Many Gower, Pembs and Vale beaches ban dogs from 1 May – 30 September (sometimes 1 Oct). Always check local signage."),
    ("", ""),
    ("Waymarked", ""),
    ("Yes", "Clearly signposted throughout (e.g. Wales Coast Path acorn, NT arrows, Cadw, NCN cycle route)."),
    ("Partial", "Some waymarking but navigation aid (OS Explorer map or GPX) recommended."),
    ("No", "Navigation by map and compass."),
    ("", ""),
    ("Pushchair Friendly", ""),
    ("Yes", "All-terrain pushchair should manage most or all of the route."),
    ("Partial", "Good for part of the route or with an off-road/jogger-style pushchair."),
    ("No", "Too rough, steep, stiled or bouldery for pushchairs."),
    ("", ""),
    ("Elevation gain", "Approximate total ascent in metres. Useful cross-reference with distance for difficulty."),
    ("Estimated time", "Based on Naismith's rule (~3mph flat + 1hr per 600m ascent) with small allowance for photo/picnic stops. Slower with children."),
    ("", ""),
    ("Key abbreviations", ""),
    ("NT", "National Trust"),
    ("Cadw", "Welsh government's historic environment service (castles, abbeys)."),
    ("RSPB", "Royal Society for the Protection of Birds."),
    ("NNR / SSSI", "National Nature Reserve / Site of Special Scientific Interest — stay on paths."),
    ("NCN", "National Cycle Network route."),
    ("MoD", "Ministry of Defence — firing ranges (Castlemartin, Pendine, Giltar/Penally) — check access days."),
    ("", ""),
    ("Important notes", ""),
    ("Verify before you go", "Parking fees, pub opening hours, ferry schedules (Caldey, Symonds Yat) and MoD firing days change. Always confirm before setting out."),
    ("Weather & conditions", "Brecon Beacons summits can have Arctic conditions in winter. Always check the Met Office mountain forecast."),
    ("Tides", "Worm's Head, Broughton, Amroth low-tide walks, and several Pembrokeshire causeways are tide-dependent. Check tide tables."),
    ("Maps", "OS Explorer OL12 (Brecon Beacons west), OL13 (Brecon Beacons east), 164 (Gower), OL36 (South Pembs), 152 (Newport/Pontypool), OL14 (Wye Valley/Forest of Dean)."),
    ("Apps", "OS Maps, Komoot, AllTrails — search by walk name for GPX tracks."),
    ("Disclaimer", "This database is a starting point compiled from general knowledge as of 2026. Always check official sources (NRW, NT, Cadw, local councils) for the latest access, parking and safety information."),
    ("", ""),
    ("New columns (v2)", ""),
    ("Start Postcode", "Approximate UK postcode for the suggested car park / walk start. Drop into any sat nav or mapping app."),
    ("Drive from Monmouth (mins)", "Approximate drive time from Monmouth NP25 3NT, in minutes. Filter by <= 60 to find walks within one hour."),
    ("Map Link", "Clickable link — opens Google Maps at the start point so you can verify location and get directions from your current location."),
]

for r, (k, v) in enumerate(legend_rows, start=1):
    c1 = legend.cell(row=r, column=1, value=k)
    c2 = legend.cell(row=r, column=2, value=v)
    c1.font = Font(name="Arial", bold=(r == 1 or v == ""), size=11 if r == 1 else 10)
    c2.font = Font(name="Arial", size=10)
    c1.alignment = Alignment(vertical="top", wrap_text=True)
    c2.alignment = Alignment(vertical="top", wrap_text=True)
    if r == 1:
        c1.font = Font(name="Arial", bold=True, size=14, color="1F4E78")
    if k and v == "" and r > 1:
        c1.font = Font(name="Arial", bold=True, size=11, color="1F4E78")
        c1.fill = PatternFill("solid", start_color="EAF1FA")
    legend.row_dimensions[r].height = 20 if len(v) < 80 else 40

# ---------------------------------------------------------------------------
# Regions summary sheet (formulas)
# ---------------------------------------------------------------------------
summary = wb.create_sheet("Region Summary")
summary.column_dimensions['A'].width = 36
for col in ['B', 'C', 'D', 'E', 'F']:
    summary.column_dimensions[col].width = 16

headers = ["Region", "Walks", "Total miles", "Avg distance", "Avg elevation (m)", "Avg time (hrs)"]
for c, h in enumerate(headers, 1):
    cell = summary.cell(row=1, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center

regions = [BRECON, GOWER, PEMBS, VALES, WYEMON, MIDWAL, CARMS, BORDERS]
for r, region in enumerate(regions, start=2):
    summary.cell(row=r, column=1, value=region).font = Font(name="Arial", size=10)
    summary.cell(row=r, column=2, value=f'=COUNTIF(Walks!C:C,A{r})').font = Font(name="Arial", size=10)
    summary.cell(row=r, column=3, value=f'=SUMIF(Walks!C:C,A{r},Walks!F:F)').font = Font(name="Arial", size=10)
    summary.cell(row=r, column=4, value=f'=IFERROR(C{r}/B{r},"")').font = Font(name="Arial", size=10)
    summary.cell(row=r, column=5, value=f'=IFERROR(SUMIF(Walks!C:C,A{r},Walks!H:H)/B{r},"")').font = Font(name="Arial", size=10)
    summary.cell(row=r, column=6, value=f'=IFERROR(SUMIF(Walks!C:C,A{r},Walks!I:I)/B{r},"")').font = Font(name="Arial", size=10)
    summary.cell(row=r, column=3).number_format = "0.0"
    summary.cell(row=r, column=4).number_format = "0.0"
    summary.cell(row=r, column=5).number_format = "0"
    summary.cell(row=r, column=6).number_format = "0.0"

# Total row
total_row = len(regions) + 2
summary.cell(row=total_row, column=1, value="All of South Wales").font = Font(name="Arial", bold=True)
summary.cell(row=total_row, column=2, value=f'=SUM(B2:B{total_row-1})').font = Font(name="Arial", bold=True)
summary.cell(row=total_row, column=3, value=f'=SUM(C2:C{total_row-1})').font = Font(name="Arial", bold=True)
summary.cell(row=total_row, column=4, value=f'=IFERROR(C{total_row}/B{total_row},"")').font = Font(name="Arial", bold=True)
summary.cell(row=total_row, column=5, value=f'=IFERROR(SUMPRODUCT(B2:B{total_row-1},E2:E{total_row-1})/B{total_row},"")').font = Font(name="Arial", bold=True)
summary.cell(row=total_row, column=6, value=f'=IFERROR(SUMPRODUCT(B2:B{total_row-1},F2:F{total_row-1})/B{total_row},"")').font = Font(name="Arial", bold=True)
summary.cell(row=total_row, column=3).number_format = "0.0"
summary.cell(row=total_row, column=4).number_format = "0.0"
summary.cell(row=total_row, column=5).number_format = "0"
summary.cell(row=total_row, column=6).number_format = "0.0"
for c in range(1, 7):
    summary.cell(row=total_row, column=c).fill = PatternFill("solid", start_color="EAF1FA")

# ---------------------------------------------------------------------------
# Near-Monmouth sheet: walks within 60 minutes' drive of NP25 3NT
# ---------------------------------------------------------------------------
near = wb.create_sheet("Near Monmouth (<60 min)")
near_cols = [
    ("ID", 6),
    ("Walk Name", 40),
    ("Region", 22),
    ("Distance (mi)", 12),
    ("Elevation (m)", 12),
    ("Est. Time (hrs)", 13),
    ("Difficulty", 12),
    ("Dogs Allowed", 14),
    ("Start Postcode", 14),
    ("Drive (mins)", 12),
    ("Map Link", 22),
]
for c, (h, w) in enumerate(near_cols, 1):
    cell = near.cell(row=1, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    near.column_dimensions[get_column_letter(c)].width = w

near_rows = []
for i, w in enumerate(WALKS, start=1):
    name = w[0]
    region = w[1]
    miles = w[4]
    elev = w[5]
    time_h = w[6]
    diff = w[7]
    dogs = w[10]
    pc, drive = POSTCODES.get(name, ("", None))
    if isinstance(drive, int) and drive <= 60:
        near_rows.append((i, name, region, miles, elev, time_h, diff, dogs, pc, drive, map_link(pc)))

near_rows.sort(key=lambda r: r[9])  # sort by drive time

for r, row in enumerate(near_rows, start=2):
    for c, v in enumerate(row, start=1):
        cell = near.cell(row=r, column=c, value=v)
        cell.font = cell_font
        cell.alignment = center if c != 2 and c != 11 else (left_wrap if c == 2 else center)
    # hyperlink
    link_cell = near.cell(row=r, column=11)
    link_cell.value = "Open in Google Maps"
    link_cell.hyperlink = row[10]
    link_cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
    near.cell(row=r, column=4).number_format = "0.0"

near.freeze_panes = "A2"
near.auto_filter.ref = f"A1:K{len(near_rows) + 1}"
near.row_dimensions[1].height = 30

# Tab colours
ws.sheet_properties.tabColor = "1F4E78"
legend.sheet_properties.tabColor = "808080"
summary.sheet_properties.tabColor = "228B22"
near.sheet_properties.tabColor = "C8102E"

from pathlib import Path
output_path = Path(__file__).parent / "South_Wales_Walks_Database.xlsx"
wb.save(output_path)
print(f"Saved {output_path}")
print(f"Walks: {len(WALKS)}")
print(f"Columns: {len(COLUMNS)}")
